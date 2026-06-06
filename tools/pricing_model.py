#!/usr/bin/env python3
"""
Pricing Model — usage-based (AUM / seat / volume) commercial pricing engine.

The price side of the table. Where `roi_excel_generator.py` models the CLIENT'S
benefit, this models what BACKBASE charges and how it scales with the client's
growth — band schedules, smoothing, crossover, margin head-to-head, and the
software-vs-3rd-party back-solve needed for a POF / BAFO submission.

Two layers (both surfaced on real deals):
  1. SCALAR scenario model — 5-yr-total × band multipliers. Drives the scenario
     projection tables and crossover analysis on negotiation decks.
  2. POF granular model — per-line per-year fees; uplift applied to SOFTWARE only;
     3rd-party (pass-through) held flat; back-solve a proportional discount on the
     uplift so the grand total hits a target. (Exactly the Phase-A POF maths.)

Usage:
    python pricing_model.py --config model.json --out model.xlsx     # build workbook
    python pricing_model.py --selftest                               # regression test
    python pricing_model.py --config model.json --crossover A,B      # print a crossover

Config schema: see `--print-schema` or knowledge/learnings/pipeline_gaps/SPEC_pricing-model.md
Origin: knowledge/learnings/pipeline_gaps/deal_desk_commercial_track.md
"""

import json
import argparse
from pathlib import Path

# ── Brand (frontline-tokens.json) ────────────────────────────────────
NAVY = "041326"
BLUE = "3367FF"
LIGHT_BLUE = "E5EBFF"
GREEN = "15803D"
LIGHT_GREEN = "D1FAE5"
GREY = "6B7280"
RED = "DC2626"


# ════════════════════════════════════════════════════════════════════
#  SCALAR SCENARIO MODEL
# ════════════════════════════════════════════════════════════════════

def scalar_price(scenario, g):
    """Fee for a scenario at growth value `g`. Dispatches on `basis`:

      - "band_multiplier" (default · WEALTH / AUM):
            baseline × Π(1+rate) for each band crossed, × top-band smoothing.
      - "tiered_per_unit" (RETAIL / BANKING / LENDING):
            flat platform fee + Σ(units in each volume tier × per-unit rate).

    Both return a fee in the scenario's own value unit; keep scenarios within
    one model internally consistent so they can be compared / crossed.
    """
    basis = scenario.get("basis", "band_multiplier")
    if basis == "tiered_per_unit":
        return price_tiered(scenario, g)
    return price_band_multiplier(scenario, g)


def price_band_multiplier(scenario, g):
    """WEALTH / AUM. A band [lo, hi, rate] multiplies the fee by (1+rate) once
    `g` crosses `lo`. The open `top_band` {from, step, rate} compounds
    (1+rate)^((g-from)/step) once `g` exceeds `from` (smoothing = lower rate)."""
    price = float(scenario["baseline"])
    for lo, hi, rate in scenario.get("bands", []):
        if g > lo:
            price *= (1.0 + rate)
    tb = scenario.get("top_band")
    if tb and g > tb["from"]:
        steps = (g - tb["from"]) / tb["step"]
        price *= (1.0 + tb["rate"]) ** steps
    return price


def price_tiered(scenario, g):
    """RETAIL / BANKING / LENDING. Flat platform fee plus a per-unit charge that
    steps down by volume tier — the standard customer/account/loan model.

      flat            : fixed platform fee (same value unit as the output)
      tiers           : list of [lo, hi, per_unit]; hi=null means open-ended.
                        Units between lo and min(g,hi) are charged at per_unit.

    Volume discounting (lower per_unit in higher tiers) is the retail analogue
    of the wealth reducing-uplift curve.
    """
    price = float(scenario.get("flat", 0.0))
    for tier in scenario.get("tiers", []):
        lo, hi, per_unit = tier[0], tier[1], tier[2]
        hi = g if hi is None else min(g, hi)
        if g > lo:
            price += (hi - lo) * per_unit
    return price


def projection(model):
    """Return {milestone: {scenario_name: price}} across all milestones."""
    sm = model["scalar_model"]
    out = {}
    for g in sm["milestones"]:
        out[g] = {s["name"]: scalar_price(s, g) for s in sm["scenarios"]}
    return out


def find_crossover(model, name_a, name_b, lo=None, hi=None, step=None):
    """Smallest growth value where scenario A and B swap which is cheaper.
    Scan range defaults to the model's milestone span, so it works for any
    metric scale (AUM £bn, customer counts, loan volumes, ...).
    Returns (crossover_value, 'A_cheaper_below'|'B_cheaper_below'|None)."""
    sm = model["scalar_model"]
    sa = next(s for s in sm["scenarios"] if s["name"] == name_a)
    sb = next(s for s in sm["scenarios"] if s["name"] == name_b)
    ms = sm["milestones"]
    if lo is None:
        lo = min(min(ms), sa.get("baseline_at", min(ms)), sb.get("baseline_at", min(ms)))
    if hi is None:
        hi = max(ms)
    if step is None:
        step = max((hi - lo) / 2000.0, 1e-9)
    prev = scalar_price(sa, lo) - scalar_price(sb, lo)
    g = lo
    while g <= hi:
        d = scalar_price(sa, g) - scalar_price(sb, g)
        if prev == 0:
            prev = d
        if d == 0 or (d > 0) != (prev > 0):
            direction = "A_cheaper_below" if prev < 0 else "B_cheaper_below"
            return round(g, 1), direction
        prev = d
        g += step
    return None, None


def margin_h2h(model, name_a, name_b):
    """Per-milestone Backbase-revenue delta between two structures.
    Positive = A yields more to Backbase than B at that milestone."""
    sm = model["scalar_model"]
    sa = next(s for s in sm["scenarios"] if s["name"] == name_a)
    sb = next(s for s in sm["scenarios"] if s["name"] == name_b)
    rows = []
    for g in sm["milestones"]:
        pa, pb = scalar_price(sa, g), scalar_price(sb, g)
        rows.append({"g": g, name_a: pa, name_b: pb, "delta_A_minus_B": pa - pb})
    return rows


# ════════════════════════════════════════════════════════════════════
#  POF GRANULAR MODEL  (software vs 3rd-party, back-solve discount)
# ════════════════════════════════════════════════════════════════════

def pof_backsolve(pof):
    """Apply uplift to SOFTWARE only; hold 3rd-party flat; back-solve a
    proportional discount on the uplift so software+uplift+3rd-party = target.

    Returns a dict with per-year breakdown, the solved discount, and totals.
    Reproduces the deal-desk POF maths exactly.
    """
    sw = [float(x) for x in pof["software_by_year"]]
    tp = [float(x) for x in pof["thirdparty_by_year"]]
    rate = float(pof["uplift_rate"])
    target = float(pof["target_total"])

    sw_total = sum(sw)
    tp_total = sum(tp)
    uplift_year = [s * rate for s in sw]          # full uplift per year
    uplift_total = sum(uplift_year)

    needed_uplift = target - sw_total - tp_total  # how much uplift the target leaves room for
    if uplift_total == 0:
        discount = 0.0
    else:
        discount = 1.0 - (needed_uplift / uplift_total)   # proportional reduction on uplift

    revised_uplift_year = [u * (1.0 - discount) for u in uplift_year]
    total_year = [sw[i] + revised_uplift_year[i] + tp[i] for i in range(len(sw))]

    return {
        "software_by_year": sw,
        "thirdparty_by_year": tp,
        "uplift_full_by_year": uplift_year,
        "revised_uplift_by_year": revised_uplift_year,
        "total_by_year": total_year,
        "software_total": sw_total,
        "thirdparty_total": tp_total,
        "uplift_full_total": uplift_total,
        "revised_uplift_total": sum(revised_uplift_year),
        "discount_on_uplift": discount,
        "grand_total": sum(total_year),
        "target_total": target,
        "uplift_rate": rate,
    }


# ════════════════════════════════════════════════════════════════════
#  XLSX OUTPUT
# ════════════════════════════════════════════════════════════════════

def build_xlsx(model, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    cur = model.get("currency", "£")
    unit = model.get("unit", "bn")
    metric = model.get("metric", "AUM")
    disp = model.get("display", {})
    div = float(disp.get("divisor", 1))   # divide fee values for display (e.g. 1e6 for absolute → M)
    suf = disp.get("suffix", "M")         # fee display suffix
    mpre = model.get("metric_prefix", cur)   # growth-axis label prefix (currency for AUM, "" for counts)
    msuf = model.get("metric_suffix", unit)  # growth-axis label suffix ("bn", " cust", ...)
    glabel = lambda g: f"{mpre}{g:,}{msuf}"  # format a growth-metric value for display
    thin = Side(border_style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr(ws, row, headers, start=1):
        for i, h in enumerate(headers):
            c = ws.cell(row=row, column=start + i, value=h)
            c.font = Font(bold=True, color="FFFFFF", size=10)
            c.fill = PatternFill("solid", fgColor=NAVY)
            c.alignment = Alignment(horizontal="center", wrap_text=True)
            c.border = border

    def money_fmt(cell):
        cell.number_format = f'"{cur}"#,##0.00,,"M"' if False else '#,##0.00'

    wb = Workbook()

    # ── Tab 1 · Scenarios projection ──────────────────────────────────
    ws = wb.active
    ws.title = "Scenarios"
    ws["A1"] = f"{model.get('client','Client')} — pricing scenarios across {metric} growth"
    ws["A1"].font = Font(bold=True, size=14, color=NAVY)
    sm = model["scalar_model"]
    names = [s["name"] for s in sm["scenarios"]]
    hdr(ws, 3, [metric] + names)
    proj = projection(model)
    r = 4
    for g in sm["milestones"]:
        ws.cell(row=r, column=1, value=glabel(g)).font = Font(bold=True, color=NAVY, size=10)
        ws.cell(row=r, column=1).border = border
        for i, nm in enumerate(names):
            c = ws.cell(row=r, column=2 + i, value=round(proj[g][nm] / div, 2))
            c.number_format = f'"{cur}"#,##0.00"{suf}"'
            c.alignment = Alignment(horizontal="right")
            c.border = border
        r += 1
    for i, w in enumerate([16] + [18] * len(names)):
        ws.column_dimensions[get_column_letter(1 + i)].width = w

    # ── Tab 2 · Crossovers ────────────────────────────────────────────
    ws2 = wb.create_sheet("Crossovers")
    ws2["A1"] = "Scenario crossovers (where the cheaper option swaps)"
    ws2["A1"].font = Font(bold=True, size=13, color=NAVY)
    hdr(ws2, 3, ["Pair", "Crossover " + metric, "Read"])
    r = 4
    for i in range(len(names)):
        for j in range(i + 1, len(names)):
            cv, direction = find_crossover(model, names[i], names[j])
            read = "no crossover in range" if cv is None else \
                f"{names[i] if 'A_cheaper' in (direction or '') else names[j]} cheaper below {glabel(cv)}"
            ws2.cell(row=r, column=1, value=f"{names[i]}  vs  {names[j]}").border = border
            ws2.cell(row=r, column=2, value=(glabel(cv) if cv else "—")).border = border
            ws2.cell(row=r, column=3, value=read).border = border
            r += 1
    for i, w in enumerate([34, 18, 46]):
        ws2.column_dimensions[get_column_letter(1 + i)].width = w

    # ── Tab 3 · Margin H2H (first two scenarios, if >=2) ───────────────
    if len(names) >= 2:
        a, b = names[-2], names[-1]  # default: compare the last two (usually the live options)
        ws3 = wb.create_sheet("Margin H2H")
        ws3["A1"] = f"Backbase revenue head-to-head — {a}  vs  {b}"
        ws3["A1"].font = Font(bold=True, size=13, color=NAVY)
        ws3["A2"] = "Positive delta = the first option yields more to Backbase at that point."
        ws3["A2"].font = Font(italic=True, size=10, color=GREY)
        hdr(ws3, 4, [metric, a, b, f"{a} − {b}", "Better for Backbase"])
        r = 5
        for row in margin_h2h(model, a, b):
            ws3.cell(row=r, column=1, value=glabel(row['g'])).font = Font(bold=True, color=NAVY)
            for col, key in [(2, a), (3, b), (4, "delta_A_minus_B")]:
                c = ws3.cell(row=r, column=col, value=round(row[key] / div, 2))
                c.number_format = f'"{cur}"#,##0.00"{suf}"'
                c.alignment = Alignment(horizontal="right")
                c.border = border
            better = a if row["delta_A_minus_B"] > 0 else (b if row["delta_A_minus_B"] < 0 else "tie")
            cc = ws3.cell(row=r, column=5, value=better)
            cc.font = Font(bold=True, color=BLUE if better == a else (GREEN if better == b else GREY))
            cc.border = border
            ws3.cell(row=r, column=1).border = border
            r += 1
        for i, w in enumerate([14, 18, 18, 16, 22]):
            ws3.column_dimensions[get_column_letter(1 + i)].width = w

    # ── Tab 4 · POF split (if present) ────────────────────────────────
    if "pof" in model:
        res = pof_backsolve(model["pof"])
        ws4 = wb.create_sheet("POF Split")
        ws4["A1"] = "Price-of-Failure split — uplift on software only, 3rd-party held flat"
        ws4["A1"].font = Font(bold=True, size=13, color=NAVY)
        ws4["A2"] = (f"Target total {cur}{res['target_total']:,.0f}  ·  uplift rate {res['uplift_rate']*100:.0f}%"
                     f"  ·  solved discount on uplift {res['discount_on_uplift']*100:.4f}%")
        ws4["A2"].font = Font(italic=True, size=10, color=GREY)
        nyr = len(res["software_by_year"])
        hdr(ws4, 4, ["Line"] + [f"Yr {i+1}" for i in range(nyr)] + ["5-Yr Total"])
        tp_label = model["pof"].get("thirdparty_label", "3rd-party (pass-through)")
        rows = [
            ("Software base fee", res["software_by_year"], res["software_total"], NAVY),
            (f"Uplift @ {res['uplift_rate']*100:.0f}% (full)", res["uplift_full_by_year"], res["uplift_full_total"], GREY),
            ("Uplift after discount", res["revised_uplift_by_year"], res["revised_uplift_total"], GREEN),
            (tp_label + " (flat)", res["thirdparty_by_year"], res["thirdparty_total"], BLUE),
            ("TOTAL", res["total_by_year"], res["grand_total"], NAVY),
        ]
        r = 5
        for label, series, tot, color in rows:
            cl = ws4.cell(row=r, column=1, value=label)
            cl.font = Font(bold=(label == "TOTAL"), color=color, size=10)
            cl.border = border
            for i, v in enumerate(series):
                c = ws4.cell(row=r, column=2 + i, value=round(v, 2))
                c.number_format = '#,##0'
                c.alignment = Alignment(horizontal="right")
                c.border = border
            ct = ws4.cell(row=r, column=2 + nyr, value=round(tot, 2))
            ct.number_format = '#,##0'
            ct.font = Font(bold=True, color=color)
            ct.alignment = Alignment(horizontal="right")
            ct.border = border
            if label == "TOTAL":
                for col in range(1, 3 + nyr):
                    ws4.cell(row=r, column=col).fill = PatternFill("solid", fgColor=LIGHT_GREEN)
            r += 1
        ws4.column_dimensions["A"].width = 30
        for i in range(nyr + 1):
            ws4.column_dimensions[get_column_letter(2 + i)].width = 14

    # ── Tab 5 · Assumptions ───────────────────────────────────────────
    wsa = wb.create_sheet("Assumptions")
    wsa["A1"] = "Assumptions & inputs (echo of the config — every number traces here)"
    wsa["A1"].font = Font(bold=True, size=13, color=NAVY)
    r = 3
    for s in sm["scenarios"]:
        basis = s.get("basis", "band_multiplier")
        wsa.cell(row=r, column=1, value=s["name"]).font = Font(bold=True, color=NAVY)
        if basis == "tiered_per_unit":
            wsa.cell(row=r, column=2, value=f"[tiered_per_unit] flat fee {cur}{s.get('flat',0):,.0f}")
            r += 1
            for tier in s.get("tiers", []):
                lo, hi, pu = tier[0], tier[1], tier[2]
                hi_lbl = "∞" if hi is None else f"{hi:,}"
                wsa.cell(row=r, column=2, value=f"{lo:,}–{hi_lbl} {metric}: {cur}{pu}/unit")
                r += 1
        else:
            wsa.cell(row=r, column=2, value=f"[band_multiplier] baseline {cur}{s['baseline']}{suf} at {cur}{s.get('baseline_at','?')}{unit}")
            r += 1
            for lo, hi, rate in s.get("bands", []):
                wsa.cell(row=r, column=2, value=f"{cur}{lo}–{hi}{unit}: +{rate*100:.0f}%")
                r += 1
            tb = s.get("top_band")
            if tb:
                wsa.cell(row=r, column=2, value=f"{cur}{tb['from']}{unit}+ per {cur}{tb['step']}{unit}: +{tb['rate']*100:.0f}%")
                r += 1
        r += 1
    wsa.column_dimensions["A"].width = 22
    wsa.column_dimensions["B"].width = 48

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


# ════════════════════════════════════════════════════════════════════
#  SELF-TEST  (regression against the real deal-desk numbers)
# ════════════════════════════════════════════════════════════════════

SCHRODERS_FIXTURE = {
    "client": "Regression", "currency": "£", "unit": "bn", "metric": "AUM",
    "scalar_model": {
        "milestones": [85, 100, 150, 200, 250, 300, 350, 400, 450, 500],
        "scenarios": [
            {"name": "26 March", "baseline": 14.19, "baseline_at": 80,
             "bands": [[80, 100, 0.08], [100, 150, 0.06], [150, 200, 0.05]],
             "top_band": {"from": 200, "step": 50, "rate": 0.05}},
            {"name": "A", "baseline": 14.19, "baseline_at": 80,
             "bands": [[80, 100, 0.08], [100, 150, 0.06], [150, 200, 0.05]],
             "top_band": {"from": 200, "step": 50, "rate": 0.04}},
            {"name": "B", "baseline": 15.45, "baseline_at": 100,
             "bands": [[100, 150, 0.06], [150, 200, 0.05]],
             "top_band": {"from": 200, "step": 50, "rate": 0.03}},
            {"name": "C", "baseline": 15.00, "baseline_at": 100,
             "bands": [[100, 150, 0.06], [150, 200, 0.05]],
             "top_band": {"from": 200, "step": 50, "rate": 0.04}},
        ],
    },
    "pof": {
        "uplift_rate": 0.08,
        "target_total": 15000000,
        "software_by_year": [1892820.4, 2597663, 3007826.5, 3007826.5, 3007826.5],
        "thirdparty_by_year": [162400, 118240, 124664, 131730, 139503],
        "thirdparty_label": "Marketplace (3rd-party)",
    },
}

# Expected scalar 5-yr totals (verified off the v8 deck projection table)
EXPECT_SCALAR = {
    "A":        {85: 15.33, 100: 15.33, 150: 16.24, 200: 17.06, 300: 18.45, 500: 21.58},
    "B":        {85: 15.45, 100: 15.45, 150: 16.38, 200: 17.20, 300: 18.24, 500: 20.53},
    "C":        {100: 15.00, 150: 15.90, 200: 16.70, 250: 17.36, 300: 18.06, 500: 21.12},
    "26 March": {500: 22.86},
}

# RETAIL / BANKING / LENDING flavour (illustrative — NOT real SparD numbers).
# Per-customer volume tiering with a flat platform fee. Demonstrates the second
# pricing basis the engine supports; numbers are hand-computed for regression.
RETAIL_FIXTURE = {
    "client": "Illustrative Retail Bank", "currency": "€", "unit": "", "metric": "customers",
    "display": {"divisor": 1_000_000, "suffix": "M"},
    "metric_prefix": "", "metric_suffix": " cust",
    "scalar_model": {
        "milestones": [50000, 100000, 300000, 500000, 600000, 1000000],
        "scenarios": [
            {"name": "Standard", "basis": "tiered_per_unit", "flat": 2_000_000,
             "tiers": [[0, 100000, 40], [100000, 500000, 30], [500000, None, 20]]},
            {"name": "Committed", "basis": "tiered_per_unit", "flat": 3_000_000,
             "tiers": [[0, 100000, 32], [100000, 500000, 24], [500000, None, 16]]},
        ],
    },
}
# Hand-computed expected absolute totals for the Standard tiered scenario:
#   50k : 2,000,000 + 50,000×40                                   = 4,000,000
#   300k: 2,000,000 + 100,000×40 + 200,000×30                    = 12,000,000
#   600k: 2,000,000 + 100,000×40 + 400,000×30 + 100,000×20       = 20,000,000
EXPECT_TIERED = {"Standard": {50000: 4_000_000, 300000: 12_000_000, 600000: 20_000_000}}


def selftest():
    ok = True
    # 1 · WEALTH / AUM — band_multiplier basis
    print("— WEALTH (band_multiplier) —")
    for scen, points in EXPECT_SCALAR.items():
        s = next(x for x in SCHRODERS_FIXTURE["scalar_model"]["scenarios"] if x["name"] == scen)
        for g, expected in points.items():
            got = round(scalar_price(s, g), 2)
            mark = "ok" if got == expected else "FAIL"
            if got != expected:
                ok = False
            print(f"  {scen:>8} @ £{g}bn = £{got:.2f}M  (expect £{expected:.2f}M)  [{mark}]")
    # 1b · RETAIL / LENDING — tiered_per_unit basis
    print("— RETAIL (tiered_per_unit) —")
    for scen, points in EXPECT_TIERED.items():
        s = next(x for x in RETAIL_FIXTURE["scalar_model"]["scenarios"] if x["name"] == scen)
        for v, expected in points.items():
            got = round(scalar_price(s, v), 2)
            mark = "ok" if got == expected else "FAIL"
            if got != expected:
                ok = False
            print(f"  {scen:>8} @ {v:,} cust = €{got:,.0f}  (expect €{expected:,.0f})  [{mark}]")
    print("— POF back-solve —")
    # 2 · POF back-solve
    res = pof_backsolve(SCHRODERS_FIXTURE["pof"])
    checks = [
        ("discount on uplift", round(res["discount_on_uplift"] * 100, 4), 25.1237),
        ("revised uplift total", round(res["revised_uplift_total"], 0), 809500.0),
        ("grand total", round(res["grand_total"], 0), 15000000.0),
    ]
    for label, got, expected in checks:
        mark = "ok" if got == expected else "FAIL"
        if got != expected:
            ok = False
        print(f"  POF {label:>22} = {got}  (expect {expected})  [{mark}]")
    print("\nSELFTEST:", "PASS ✓" if ok else "FAIL ✗")
    return ok


# ════════════════════════════════════════════════════════════════════
def main():
    ap = argparse.ArgumentParser(description="Usage-based commercial pricing engine")
    ap.add_argument("--config", help="JSON pricing config")
    ap.add_argument("--out", help="output xlsx path")
    ap.add_argument("--crossover", help="two scenario names, comma-separated (e.g. A,B)")
    ap.add_argument("--selftest", action="store_true", help="run regression test")
    ap.add_argument("--print-schema", action="store_true")
    args = ap.parse_args()

    if args.selftest:
        raise SystemExit(0 if selftest() else 1)
    if args.print_schema:
        print(json.dumps(SCHRODERS_FIXTURE, indent=2))
        return
    if not args.config:
        ap.error("provide --config (or --selftest)")

    model = json.loads(Path(args.config).read_text())
    if args.crossover:
        a, b = args.crossover.split(",")
        cv, d = find_crossover(model, a.strip(), b.strip())
        print(f"Crossover {a}/{b}: {cv}  ({d})")
        return
    out = args.out or "pricing_model.xlsx"
    path = build_xlsx(model, out)
    print(f"Wrote {path}")


if __name__ == "__main__":
    main()
