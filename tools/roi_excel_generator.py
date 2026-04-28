#!/usr/bin/env python3
"""
ROI Excel Model Generator — Formula-Based Architecture

Generates Backbase ROI models with LIVE Excel formulas:
- Input → Calculation → Output sheet architecture
- Named ranges for global parameters
- Cross-sheet formula references (no hardcoded values in calculation cells)
- Scenario switching via dropdown (INDEX/MATCH, no VBA)
- Supports both old 'journeys' and new 'value_lever_groups' config formats

Usage:
    python roi_excel_generator.py --output output.xlsx --config roi_config.json
"""

import json
import argparse
import re
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("openpyxl not installed. Run: pip install openpyxl")
    exit(1)


class ROIModelGenerator:
    """Generates formula-based ROI Excel models following Backbase methodology."""

    COLORS = {
        'primary': '1A56FF',
        'dark': '1A1F36',
        'header_bg': '2D3250',
        'positive': '00B386',
        'negative': 'FF4444',
        'highlight': 'FFF3CD',
        'light_blue': 'E8F0FF',
        'light_gray': 'F5F5F5',
        'transcript': 'D4EDDA',
        'estimate': 'FFF3CD',
        'benchmark': 'CCE5FF',
        'low_confidence': 'F8D7DA',
        'editable': 'E8F0FF',
    }

    def __init__(self, config: dict):
        self.config = config
        self.wb = Workbook()
        self.styles_created = False
        self.cell_map = {}
        self.lever_groups = {}
        self.group_order = []
        self.lever_sheet_names = {}
        self.sheet_index = 0

    # ── Config Normalization ──────────────────────────────────────────

    def _normalize_config(self):
        """Transform journeys or value_lever_groups into unified internal format."""
        if 'value_lever_groups' in self.config:
            self.lever_groups = self.config['value_lever_groups']
        elif 'journeys' in self.config:
            self.lever_groups = {}
            for key, journey in self.config['journeys'].items():
                group = dict(journey)
                group.setdefault('group_name', key.replace('_', ' ').title())
                group.setdefault('category', self._infer_category(key))
                group.setdefault('lifecycle_stage', self._infer_lifecycle(group['category']))
                totals = group.get('totals', {})
                if 'journey_total' in totals and 'group_total' not in totals:
                    totals['group_total'] = totals['journey_total']
                self.lever_groups[key] = group
        else:
            self.lever_groups = {}

        self.group_order = list(self.lever_groups.keys())

        # Normalize backbase_loading from flat arrays to per-category curves dict
        bl = self.config.get('backbase_loading', {})
        if 'implementation_curve' in bl and 'curves' not in bl:
            bl['curves'] = {
                'default': {
                    'implementation': bl['implementation_curve'],
                    'effectiveness': bl.get('effectiveness_curve', [0.15, 0.35, 0.6, 0.85, 1.0])
                }
            }

        # Normalize scenarios from flat arrays to per-category curves
        for sc_name, sc in self.config.get('scenarios', {}).items():
            if 'implementation_curve' in sc and 'curves' not in sc:
                sc['curves'] = {
                    'default': {
                        'implementation': sc['implementation_curve'],
                        'effectiveness': sc.get('effectiveness_curve', [0.15, 0.35, 0.6, 0.85, 1.0])
                    }
                }

    @staticmethod
    def _infer_category(key):
        key_lower = key.lower()
        if 'onboard' in key_lower:
            return 'customer_onboarding'
        if 'servic' in key_lower:
            return 'servicing'
        if 'retent' in key_lower or 'churn' in key_lower or 'loyalty' in key_lower:
            return 'loyalty_retention'
        if 'cross' in key_lower or 'penetr' in key_lower or 'upsell' in key_lower:
            return 'product_penetration'
        if 'rm' in key_lower or 'productiv' in key_lower:
            return 'rm_productivity'
        return 'other'

    @staticmethod
    def _infer_lifecycle(category):
        mapping = {
            'customer_onboarding': 'acquire',
            'servicing': 'retain',
            'loyalty_retention': 'retain',
            'product_penetration': 'expand',
            'rm_productivity': 'retain',
        }
        return mapping.get(category, 'other')

    # ── Styles ────────────────────────────────────────────────────────

    def _create_styles(self):
        if self.styles_created:
            return
        styles = [
            ("header_style", Font(bold=True, color="FFFFFF", size=11),
             PatternFill("solid", fgColor=self.COLORS['header_bg']),
             Alignment(horizontal="center", vertical="center")),
            ("title_style", Font(bold=True, color=self.COLORS['primary'], size=16), None, None),
            ("input_style", None, PatternFill("solid", fgColor=self.COLORS['editable']), None),
        ]
        for name, font, fill, align in styles:
            style = NamedStyle(name=name)
            if font:
                style.font = font
            if fill:
                style.fill = fill
            if align:
                style.alignment = align
            try:
                self.wb.add_named_style(style)
            except ValueError:
                pass
        self.styles_created = True

    # ── Helpers ───────────────────────────────────────────────────────

    def _apply_confidence_coloring(self, ws, cell_ref, input_data):
        """Apply fill color based on source/confidence."""
        cell = ws[cell_ref] if isinstance(cell_ref, str) else cell_ref
        source = input_data.get('source', '')
        confidence = input_data.get('confidence', 'LOW')
        if confidence == 'HIGH' or 'TRANSCRIPT' in source.upper():
            cell.fill = PatternFill("solid", fgColor=self.COLORS['transcript'])
        elif 'BACKBASE' in source.upper():
            cell.fill = PatternFill("solid", fgColor=self.COLORS['benchmark'])
        else:
            cell.fill = PatternFill("solid", fgColor=self.COLORS['estimate'])

    def _write_header_row(self, ws, row, headers, start_col=2):
        """Write a styled header row."""
        for i, header in enumerate(headers):
            cell = ws.cell(row=row, column=start_col + i, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill("solid", fgColor=self.COLORS['header_bg'])
            cell.alignment = Alignment(horizontal="center")

    def _safe_sheet_name(self, name, prefix="Lever"):
        """Create a valid Excel sheet name (max 31 chars)."""
        clean = re.sub(r'[^\w\s\-]', '', name)
        clean = re.sub(r'\s+', ' ', clean).strip()[:22].strip()
        return f"{prefix} - {clean}"

    def _get_curve_category(self, group_key):
        """Map a lever group to its curve category in scenarios."""
        group = self.lever_groups.get(group_key, {})
        cat = group.get('category', 'default')
        # Map common categories to curve keys
        mapping = {
            'customer_onboarding': 'acquisition',
            'servicing': 'servicing',
            'loyalty_retention': 'retention',
            'product_penetration': 'penetration',
            'rm_productivity': 'rm_productivity',
        }
        curve_cat = mapping.get(cat, cat)
        # Verify it exists in scenarios, fall back to first available
        scenarios = self.config.get('scenarios', {})
        moderate = scenarios.get('moderate', scenarios.get(list(scenarios.keys())[0], {})) if scenarios else {}
        curves = moderate.get('curves', {})
        if curve_cat not in curves:
            if 'default' in curves:
                return 'default'
            if curves:
                return list(curves.keys())[0]
        return curve_cat

    # ── Config Validation ─────────────────────────────────────────────

    MAX_BACKBASE_IMPACT = 0.60
    FALLBACK_BACKBASE_IMPACT = 0.30

    def _validate_and_cap_impacts(self):
        """Cap backbase_impact values at MAX and warn on anomalies."""
        warnings = []
        total_benefit = 0
        client_revenue = self.config.get('bank_profile', {}).get('total_revenue', 0)

        for group_key, group in self.lever_groups.items():
            for driver_type in ('revenue_drivers', 'cost_drivers'):
                for drv_key, driver in group.get(driver_type, {}).items():
                    inputs = driver.get('inputs', {})
                    bi = inputs.get('backbase_impact', {})
                    val = bi.get('value', 0)
                    if val > self.MAX_BACKBASE_IMPACT:
                        old_val = val
                        bi['value'] = self.MAX_BACKBASE_IMPACT
                        warnings.append(
                            f"  ⚠ {drv_key}: backbase_impact {old_val:.0%} → capped to {self.MAX_BACKBASE_IMPACT:.0%}"
                        )
                    baseline = driver.get('baseline_annual', 0)
                    benefit = baseline * bi.get('value', self.FALLBACK_BACKBASE_IMPACT)
                    total_benefit += benefit
                    if client_revenue > 0 and benefit > client_revenue * 0.02:
                        warnings.append(
                            f"  ⚠ {drv_key}: annual benefit ${benefit:,.0f} exceeds 2% of client revenue — review baseline"
                        )

        # Also cap scenario-level backbase_impacts
        for sc_name, sc in self.config.get('scenarios', {}).items():
            for imp_key, imp_val in sc.get('backbase_impacts', {}).items():
                if isinstance(imp_val, (int, float)) and imp_val > self.MAX_BACKBASE_IMPACT:
                    old_val = imp_val
                    sc['backbase_impacts'][imp_key] = self.MAX_BACKBASE_IMPACT
                    warnings.append(
                        f"  ⚠ Scenario '{sc_name}' impact '{imp_key}': {old_val:.0%} → capped to {self.MAX_BACKBASE_IMPACT:.0%}"
                    )

        investment = self.config.get('total_investment', 0)
        if investment > 0 and total_benefit > 0:
            five_yr_roi = (total_benefit * 5 - investment) / investment * 100
            if five_yr_roi > 500:
                warnings.append(
                    f"  ⚠ REASONABLENESS CHECK: 5-year ROI = {five_yr_roi:.0f}% — review all baselines and impacts"
                )

        if warnings:
            print("ROI Validation Warnings:")
            for w in warnings:
                print(w)

        return warnings

    # ── Main Generate ─────────────────────────────────────────────────

    def generate(self, output_path: str):
        """Generate the complete formula-based ROI Excel model."""
        self._normalize_config()
        self._validate_and_cap_impacts()

        # Pre-declare sc_cell — Cashflows sheet places the dropdown at exactly C5.
        # This allows Model Inputs and Journey Analysis (written before Cashflows)
        # to reference the Cashflows scenario selector without a two-pass approach.
        self.sc_cell = "'Cashflows'!$C$5"

        self._create_styles()

        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']

        self.sheet_index = 0

        # Phase 1: Static sheets
        self._create_cover_sheet()
        self._create_instructions_sheet()
        self._create_table_of_contents()

        # Phase 2: Input sheets (needed first — other sheets reference these)
        self._create_scenario_data_sheet()
        self._create_model_inputs_sheet()

        # Phase 3: Calculation sheets
        if self._has_servicing():
            self._create_servicing_detail_sheet()
        self._create_journey_analysis_sheet()

        # Phase 4: Cashflows (includes financial metrics — merged with former Results Dashboard)
        self._create_cashflows_sheet()

        # Phase 5: Static data sheets
        if self.config.get('bank_profile'):
            self._create_bank_profile_sheet()
        self._create_assumptions_sheet()
        self._create_data_gaps_sheet()

        # Phase 6: Named ranges
        self._define_named_ranges()

        # Phase 7: Reorder sheets to match target layout
        self._reorder_sheets()


        self.wb.save(output_path)
        print(f"ROI Model saved to: {output_path}")
        return output_path

    def _reorder_sheets(self):
        """Reorder sheets: Cover, Instructions, ToC, Cashflows, Journey Analysis,
        Servicing, Model Inputs, Scenario Data, Assumptions, Data Gaps."""
        desired_order = [
            'Cover Page',
            'Instructions',
            'Table of Contents',
            'Cashflows',
            'Journey Analysis',
            'Servicing Detail',
            'Model Inputs',
            'Scenario Data',
            'Bank Profile',
            'Assumptions',
            'Data Gaps',
        ]
        # Build ordered list, skipping sheets that don't exist
        existing = list(self.wb.sheetnames)
        ordered = [name for name in desired_order if name in existing]
        # Append any remaining sheets not in the desired order
        for name in existing:
            if name not in ordered:
                ordered.append(name)
        self.wb._sheets = [self.wb[name] for name in ordered]


    def _has_servicing(self):
        """Check if any lever group has servicing_analysis."""
        return any('servicing_analysis' in g for g in self.lever_groups.values())

    # ── Cover Page (unchanged) ────────────────────────────────────────

    def _create_cover_sheet(self):
        ws = self.wb.create_sheet("Cover Page", self.sheet_index)
        self.sheet_index += 1
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 40

        ws.merge_cells('B5:C5')
        ws['B5'] = "Return on Investment Calculator"
        ws['B5'].font = Font(bold=True, size=24, color=self.COLORS['primary'])

        ws.merge_cells('B7:C7')
        ws['B7'] = "Value Consulting - Business Case"
        ws['B7'].font = Font(size=14, italic=True)

        scenario_map = {'1': 'Conservative', '2': 'Moderate', '3': 'Aggressive'}
        selected = self.config.get('selected_scenario', 'Moderate')
        info = [
            (10, "Client Name:", self.config.get('client_name', 'Client')),
            (12, "Date:", self.config.get('date', datetime.now().strftime('%Y-%m-%d'))),
            (14, "Currency:", self.config.get('currency', 'USD')),
            (16, "Analysis Period:", f"{self.config.get('analysis_years', 5)} Years"),
            (18, "Selected Scenario:", selected),
            (20, "Discount Rate:", f"{self.config.get('discount_rate', 0.10) * 100}%"),
        ]
        for row, label, value in info:
            ws[f'B{row}'] = label
            ws[f'B{row}'].font = Font(bold=True)
            ws[f'C{row}'] = value

        ws['B24'] = "DATA SOURCE LEGEND"
        ws['B24'].font = Font(bold=True, size=12)
        legends = [
            (26, "Green", "From transcript/client data", 'transcript'),
            (27, "Yellow", "Estimate - requires validation", 'estimate'),
            (28, "Blue", "Backbase benchmark", 'benchmark'),
            (29, "Red border", "Low confidence - critical validation", 'low_confidence'),
        ]
        for row, color, desc, color_key in legends:
            ws[f'B{row}'] = color
            ws[f'B{row}'].fill = PatternFill("solid", fgColor=self.COLORS[color_key])
            ws[f'C{row}'] = desc

    # ── Instructions (updated for formula model) ──────────────────────

    def _create_instructions_sheet(self):
        ws = self.wb.create_sheet("Instructions", self.sheet_index)
        self.sheet_index += 1
        ws.column_dimensions['B'].width = 80

        ws['B2'] = "ROI Model Instructions"
        ws['B2'].font = Font(bold=True, size=18, color=self.COLORS['primary'])

        instructions = [
            "",
            "THIS IS A LIVE MODEL — all calculations use Excel formulas.",
            "Changing any input on 'Model Inputs' will automatically recalculate all sheets.",
            "",
            "SCENARIO SWITCHING",
            "• Go to 'Model Inputs' sheet, cell C5",
            "• Select: 1 = Conservative, 2 = Moderate, 3 = Aggressive",
            "• All curves, impacts, and downstream calculations will update instantly",
            "",
            "EDITING INPUTS",
            "• Blue-shaded cells on 'Model Inputs' are editable — change volumes, rates, costs",
            "• Yellow cells are estimates that should be validated with client data",
            "• Green cells are from transcript — highest confidence",
            "",
            "SHEET ARCHITECTURE",
            "• Model Inputs → all editable parameters (INPUT layer)",
            "• Scenario Data → three scenario definitions (reference table)",
            "• Lever sheets → per-group calculations with formulas (CALCULATION layer)",
            "• Cashflows → 5-year projection aggregating lever sheets (CALCULATION layer)",
            "• Results Dashboard → summary KPIs from Cashflows (OUTPUT layer)",
            "",
            "DATA CONFIDENCE LEVELS",
            "• HIGH - Direct quote from transcript or validated client data",
            "• MEDIUM - Inferred from transcript context or industry benchmark",
            "• LOW - Estimate based on typical patterns - REQUIRES VALIDATION",
        ]
        for i, line in enumerate(instructions, start=4):
            ws[f'B{i}'] = line
            if line in ["SCENARIO SWITCHING", "EDITING INPUTS", "SHEET ARCHITECTURE",
                        "DATA CONFIDENCE LEVELS",
                        "THIS IS A LIVE MODEL — all calculations use Excel formulas."]:
                ws[f'B{i}'].font = Font(bold=True, color=self.COLORS['primary'])

    # ── Table of Contents (updated) ───────────────────────────────────

    def _create_table_of_contents(self):
        ws = self.wb.create_sheet("Table of Contents", self.sheet_index)
        self.sheet_index += 1
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 50

        ws['B3'] = "Table of Contents"
        ws['B3'].font = Font(bold=True, size=18, color=self.COLORS['primary'])

        sections = [
            ("Input Layer", "Model Inputs — editable parameters"),
            ("", "  Scenario Data — Conservative / Moderate / Aggressive"),
        ]
        if self._has_servicing():
            sections.append(("Calculation", "  Servicing Detail — task-level breakdown"))
        for group_key in self.group_order:
            group = self.lever_groups[group_key]
            sections.append(("Calculation" if not self._has_servicing() else "", f"  Lever: {group.get('group_name', group_key)}"))
        sections += [
            ("", "  Cashflows — 5-year projection"),
            ("Output Layer", "Results Dashboard — NPV, IRR, Payback, ROI"),
            ("Reference", "Assumptions Register"),
            ("", "  Data Gaps for Validation"),
        ]
        for i, (section, desc) in enumerate(sections, start=5):
            ws[f'B{i}'] = section
            ws[f'B{i}'].font = Font(bold=True) if section else Font()
            ws[f'C{i}'] = desc

    # ── Scenario Data Sheet ───────────────────────────────────────────

    def _create_scenario_data_sheet(self):
        ws = self.wb.create_sheet("Scenario Data", self.sheet_index)
        self.sheet_index += 1

        for col, w in [('A', 5), ('B', 35), ('C', 18), ('D', 18), ('E', 18)]:
            ws.column_dimensions[col].width = w

        ws['B2'] = "Scenario Definitions"
        ws['B2'].font = Font(bold=True, size=18, color=self.COLORS['primary'])

        ws['B3'] = "Edit values here to customize scenarios. All lever sheets reference this table."
        ws['B3'].font = Font(italic=True, size=10)

        # Headers
        self._write_header_row(ws, 5, ['Parameter', 'Conservative', 'Moderate', 'Aggressive'])

        scenarios = self.config.get('scenarios', {})
        sc_names = ['conservative', 'moderate', 'aggressive']
        sc_data = [scenarios.get(n, {}) for n in sc_names]

        # Collect all curve categories across all scenarios
        all_curve_cats = set()
        for sc in sc_data:
            all_curve_cats.update(sc.get('curves', {}).keys())
        all_curve_cats = sorted(all_curve_cats)

        self.cell_map['scenario_data'] = {'curves': {}, 'impacts': {}}
        row = 7

        # Implementation Curves
        ws[f'B{row}'] = "IMPLEMENTATION CURVES"
        ws[f'B{row}'].font = Font(bold=True, size=12, color=self.COLORS['primary'])
        row += 1

        for cat in all_curve_cats:
            ws[f'B{row}'] = cat.replace('_', ' ').title()
            ws[f'B{row}'].font = Font(bold=True)
            self.cell_map['scenario_data']['curves'].setdefault(cat, {})
            row += 1
            impl_start_row = row
            for yr in range(5):
                ws[f'B{row}'] = f"  Year {yr + 1}"
                for sc_idx, sc in enumerate(sc_data):
                    val = sc.get('curves', {}).get(cat, {}).get('implementation', [0]*5)
                    ws.cell(row=row, column=3 + sc_idx, value=val[yr] if yr < len(val) else 0)
                    ws.cell(row=row, column=3 + sc_idx).number_format = '0%'
                    ws.cell(row=row, column=3 + sc_idx).fill = PatternFill("solid", fgColor=self.COLORS['editable'])
                row += 1
            self.cell_map['scenario_data']['curves'][cat]['impl_start'] = impl_start_row
            row += 1  # blank separator

        # Effectiveness Curves
        ws[f'B{row}'] = "EFFECTIVENESS CURVES"
        ws[f'B{row}'].font = Font(bold=True, size=12, color=self.COLORS['primary'])
        row += 1

        for cat in all_curve_cats:
            ws[f'B{row}'] = cat.replace('_', ' ').title()
            ws[f'B{row}'].font = Font(bold=True)
            row += 1
            eff_start_row = row
            for yr in range(5):
                ws[f'B{row}'] = f"  Year {yr + 1}"
                for sc_idx, sc in enumerate(sc_data):
                    val = sc.get('curves', {}).get(cat, {}).get('effectiveness', [0]*5)
                    ws.cell(row=row, column=3 + sc_idx, value=val[yr] if yr < len(val) else 0)
                    ws.cell(row=row, column=3 + sc_idx).number_format = '0%'
                    ws.cell(row=row, column=3 + sc_idx).fill = PatternFill("solid", fgColor=self.COLORS['editable'])
                row += 1
            self.cell_map['scenario_data']['curves'][cat]['eff_start'] = eff_start_row
            row += 1

        # Backbase Impacts
        ws[f'B{row}'] = "BACKBASE IMPACTS"
        ws[f'B{row}'].font = Font(bold=True, size=12, color=self.COLORS['primary'])
        row += 1

        all_impact_keys = set()
        for sc in sc_data:
            all_impact_keys.update(sc.get('backbase_impacts', {}).keys())
        all_impact_keys = sorted(all_impact_keys)

        for impact_key in all_impact_keys:
            ws[f'B{row}'] = impact_key.replace('_', ' ').title()
            for sc_idx, sc in enumerate(sc_data):
                val = sc.get('backbase_impacts', {}).get(impact_key, 0)
                ws.cell(row=row, column=3 + sc_idx, value=val)
                ws.cell(row=row, column=3 + sc_idx).number_format = '0.0%'
                ws.cell(row=row, column=3 + sc_idx).fill = PatternFill("solid", fgColor=self.COLORS['editable'])
            self.cell_map['scenario_data']['impacts'][impact_key] = row
            row += 1

    # ── Model Inputs Sheet ────────────────────────────────────────────

    def _create_model_inputs_sheet(self):
        ws = self.wb.create_sheet("Model Inputs", self.sheet_index)
        self.sheet_index += 1

        for col, w in [('A', 5), ('B', 40), ('C', 20), ('D', 12), ('E', 45), ('F', 25)]:
            ws.column_dimensions[col].width = w

        self.cell_map['model_inputs'] = {'groups': {}, 'investment': {}, 'curves': {}, 'impacts': {}}

        ws['B2'] = "Model Inputs"
        ws['B2'].font = Font(bold=True, size=18, color=self.COLORS['primary'])
        ws['B3'] = "Blue cells are editable. Change values and the model recalculates."
        ws['B3'].font = Font(italic=True, size=10)

        row = 5
        # Active Scenario — read-only label referencing Cashflows dropdown
        ws[f'B{row}'] = "Active Scenario"
        ws[f'B{row}'].font = Font(bold=True, size=11, color=self.COLORS['primary'])
        sc_label_formula = (
            f"=IF({self.sc_cell}=1,\"Conservative\","
            f"IF({self.sc_cell}=2,\"Moderate\",\"Aggressive\"))"
        )
        ws.cell(row=row, column=3, value=sc_label_formula)
        ws.cell(row=row, column=3).font = Font(bold=True, size=11, color=self.COLORS['positive'])
        ws[f'D{row}'] = "(Change scenario on Cashflows sheet)"
        ws[f'D{row}'].font = Font(italic=True, size=9)

        row += 2
        # Basic Information
        ws[f'B{row}'] = "BASIC INFORMATION"
        ws[f'B{row}'].font = Font(bold=True, size=12, color=self.COLORS['primary'])
        row += 1
        self._write_header_row(ws, row, ['Input', 'Value', 'Conf.', 'Assumption/Comment', 'Source'])
        row += 1

        basic = self.config.get('basic_information', {})
        basic_start = row
        # Write all basic info fields (skip source fields)
        basic_fields = [(k, v) for k, v in basic.items() if not k.endswith('_source')]
        self.cell_map['model_inputs']['basic'] = {}
        for field_name, field_val in basic_fields:
            ws.cell(row=row, column=2, value=field_name.replace('_', ' ').title())
            ws.cell(row=row, column=3, value=field_val)
            ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor=self.COLORS['editable'])
            if isinstance(field_val, (int, float)) and field_val > 1000:
                ws.cell(row=row, column=3).number_format = '#,##0'
            elif isinstance(field_val, float) and field_val < 1:
                ws.cell(row=row, column=3).number_format = '0.00%'
            source = basic.get(f'{field_name}_source', basic.get(f'{field_name.rstrip("_annual")}_source', ''))
            conf = 'HIGH' if 'TRANSCRIPT' in str(source).upper() else ('MED' if 'BACKBASE' in str(source).upper() else 'LOW')
            ws.cell(row=row, column=4, value=conf)
            ws.cell(row=row, column=5, value=str(source)[:80])
            ws.cell(row=row, column=5).font = Font(size=9)
            self.cell_map['model_inputs']['basic'][field_name] = f'C{row}'
            row += 1

        row += 1
        # Discount Rate
        ws[f'B{row}'] = "FINANCIAL PARAMETERS"
        ws[f'B{row}'].font = Font(bold=True, size=12, color=self.COLORS['primary'])
        row += 1
        ws.cell(row=row, column=2, value="Discount Rate (WACC)")
        ws.cell(row=row, column=3, value=self.config.get('discount_rate', 0.10))
        ws.cell(row=row, column=3).number_format = '0.0%'
        ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor=self.COLORS['editable'])
        self.cell_map['model_inputs']['discount_rate'] = f'C{row}'
        row += 1
        ws.cell(row=row, column=2, value="Analysis Period (Years)")
        ws.cell(row=row, column=3, value=self.config.get('analysis_years', 5))
        row += 1
        yoy = self.config.get('backbase_loading', {}).get('yoy_growth', [0.08]*5)
        ws.cell(row=row, column=2, value="YoY Growth Rate")
        ws.cell(row=row, column=3, value=yoy[0] if yoy else 0.08)
        ws.cell(row=row, column=3).number_format = '0.0%'
        ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor=self.COLORS['editable'])
        self.cell_map['model_inputs']['yoy_growth'] = f'C{row}'

        # Lever Group Inputs
        row += 2
        ws[f'B{row}'] = "LEVER GROUP INPUTS"
        ws[f'B{row}'].font = Font(bold=True, size=12, color=self.COLORS['primary'])
        row += 1

        for group_key in self.group_order:
            group = self.lever_groups[group_key]
            self.cell_map['model_inputs']['groups'][group_key] = {'inputs': {}, 'drivers': {}}

            ws[f'B{row}'] = group.get('group_name', group_key)
            ws[f'B{row}'].font = Font(bold=True, size=11, color=self.COLORS['primary'])
            row += 1

            # Revenue drivers
            for drv_key, driver in group.get('revenue_drivers', {}).items():
                ws[f'B{row}'] = f"  {driver.get('name', drv_key)}"
                ws[f'B{row}'].font = Font(bold=True, size=10)
                row += 1
                driver_inputs = {}
                for inp_key, inp_data in driver.get('inputs', {}).items():
                    ws.cell(row=row, column=2, value=f"    {inp_key.replace('_', ' ').title()}")
                    if inp_key == 'backbase_impact':
                        # Find matching impact key and reference it
                        impact_ref = self._find_impact_ref(inp_data, driver, group_key)
                        if impact_ref:
                            # INDEX formula string → just prepend '='; cell ref → prefix sheet name
                            if impact_ref.startswith('INDEX('):
                                ws.cell(row=row, column=3, value=f"={impact_ref}")
                            else:
                                ws.cell(row=row, column=3, value=f"='Model Inputs'!{impact_ref}")
                        else:
                            ws.cell(row=row, column=3, value=inp_data.get('value', 0))
                            ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor=self.COLORS['editable'])
                    else:
                        ws.cell(row=row, column=3, value=inp_data.get('value', 0))
                        ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor=self.COLORS['editable'])
                    self._apply_confidence_coloring(ws, ws.cell(row=row, column=3), inp_data)
                    ws.cell(row=row, column=4, value=inp_data.get('confidence', 'LOW'))
                    ws.cell(row=row, column=5, value=str(inp_data.get('assumption', '')))
                    ws.cell(row=row, column=5).font = Font(size=9)
                    ws.cell(row=row, column=6, value=str(inp_data.get('source', '')))
                    ws.cell(row=row, column=6).font = Font(size=9)
                    driver_inputs[inp_key] = f'C{row}'
                    self.cell_map['model_inputs']['groups'][group_key]['inputs'][f'{drv_key}__{inp_key}'] = f'C{row}'
                    row += 1

                # Baseline Annual — formula if template exists, else hardcoded
                ws.cell(row=row, column=2, value="    Baseline Annual")
                formula_tpl = driver.get('baseline_formula', '')
                if formula_tpl:
                    excel_formula = '=' + formula_tpl
                    for key, cell_ref in driver_inputs.items():
                        excel_formula = excel_formula.replace(f'{{{key}}}', cell_ref)
                    ws.cell(row=row, column=3, value=excel_formula)
                else:
                    ws.cell(row=row, column=3, value=driver.get('baseline_annual', 0))
                ws.cell(row=row, column=3).number_format = '#,##0'
                baseline_cell = f'C{row}'
                self.cell_map['model_inputs']['groups'][group_key]['drivers'][f'{drv_key}__baseline'] = baseline_cell
                row += 1

                # Annual Benefit (Steady State) — Baseline × backbase_impact
                ws.cell(row=row, column=2, value="    Annual Benefit (Steady State)")
                backbase_cell = driver_inputs.get('backbase_impact')
                if backbase_cell and formula_tpl:
                    ws.cell(row=row, column=3, value=f'={baseline_cell}*{backbase_cell}')
                else:
                    ws.cell(row=row, column=3, value=driver.get('potential_annual_benefit', 0))
                ws.cell(row=row, column=3).number_format = '#,##0'
                ws.cell(row=row, column=3).font = Font(bold=True, color=self.COLORS['positive'])
                self.cell_map['model_inputs']['groups'][group_key]['drivers'][f'{drv_key}__benefit'] = f'C{row}'
                row += 1
                row += 1  # space between drivers

            # Cost drivers
            for drv_key, driver in group.get('cost_drivers', {}).items():
                ws[f'B{row}'] = f"  {driver.get('name', drv_key)}"
                ws[f'B{row}'].font = Font(bold=True, size=10)
                row += 1
                driver_inputs = {}
                for inp_key, inp_data in driver.get('inputs', {}).items():
                    ws.cell(row=row, column=2, value=f"    {inp_key.replace('_', ' ').title()}")
                    if inp_key == 'backbase_impact':
                        impact_ref = self._find_impact_ref(inp_data, driver, group_key)
                        if impact_ref:
                            if impact_ref.startswith('INDEX('):
                                ws.cell(row=row, column=3, value=f"={impact_ref}")
                            else:
                                ws.cell(row=row, column=3, value=f"='Model Inputs'!{impact_ref}")
                        else:
                            ws.cell(row=row, column=3, value=inp_data.get('value', 0))
                            ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor=self.COLORS['editable'])
                    else:
                        ws.cell(row=row, column=3, value=inp_data.get('value', 0))
                        ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor=self.COLORS['editable'])
                    self._apply_confidence_coloring(ws, ws.cell(row=row, column=3), inp_data)
                    ws.cell(row=row, column=4, value=inp_data.get('confidence', 'LOW'))
                    ws.cell(row=row, column=5, value=str(inp_data.get('assumption', '')))
                    ws.cell(row=row, column=5).font = Font(size=9)
                    ws.cell(row=row, column=6, value=str(inp_data.get('source', '')))
                    ws.cell(row=row, column=6).font = Font(size=9)
                    driver_inputs[inp_key] = f'C{row}'
                    self.cell_map['model_inputs']['groups'][group_key]['inputs'][f'{drv_key}__{inp_key}'] = f'C{row}'
                    row += 1

                # Baseline Annual — formula if template exists, else hardcoded
                ws.cell(row=row, column=2, value="    Baseline Annual")
                formula_tpl = driver.get('baseline_formula', '')
                if formula_tpl:
                    excel_formula = '=' + formula_tpl
                    for key, cell_ref in driver_inputs.items():
                        excel_formula = excel_formula.replace(f'{{{key}}}', cell_ref)
                    ws.cell(row=row, column=3, value=excel_formula)
                else:
                    ws.cell(row=row, column=3, value=driver.get('baseline_annual', 0))
                ws.cell(row=row, column=3).number_format = '#,##0'
                baseline_cell = f'C{row}'
                self.cell_map['model_inputs']['groups'][group_key]['drivers'][f'{drv_key}__baseline'] = baseline_cell
                row += 1

                # Annual Benefit (Steady State) — Baseline × backbase_impact
                ws.cell(row=row, column=2, value="    Annual Benefit (Steady State)")
                backbase_cell = driver_inputs.get('backbase_impact')
                if backbase_cell and formula_tpl:
                    ws.cell(row=row, column=3, value=f'={baseline_cell}*{backbase_cell}')
                else:
                    ws.cell(row=row, column=3, value=driver.get('potential_annual_benefit', 0))
                ws.cell(row=row, column=3).number_format = '#,##0'
                ws.cell(row=row, column=3).font = Font(bold=True, color=self.COLORS['positive'])
                self.cell_map['model_inputs']['groups'][group_key]['drivers'][f'{drv_key}__benefit'] = f'C{row}'
                row += 1
                row += 1

            row += 1  # space between groups

        # Active Backbase Impacts (INDEX from Scenario Data)
        row += 1
        ws[f'B{row}'] = "ACTIVE BACKBASE IMPACTS (auto-selected from scenario)"
        ws[f'B{row}'].font = Font(bold=True, size=12, color=self.COLORS['primary'])
        row += 1

        sd_impacts = self.cell_map['scenario_data']['impacts']
        for impact_key in sorted(sd_impacts.keys()):
            ws.cell(row=row, column=2, value=impact_key.replace('_', ' ').title())
            sd_row = sd_impacts[impact_key]
            formula = f"=INDEX('Scenario Data'!$C${sd_row}:$E${sd_row},1,{self.sc_cell})"
            ws.cell(row=row, column=3, value=formula)
            ws.cell(row=row, column=3).number_format = '0.0%'
            self.cell_map['model_inputs']['impacts'][impact_key] = f'C{row}'
            row += 1

        # Investment
        row += 2
        ws[f'B{row}'] = "INVESTMENT"
        ws[f'B{row}'].font = Font(bold=True, size=12, color=self.COLORS['primary'])
        row += 1

        investment = self.config.get('investment', {})
        license_data = investment.get('license', {})
        impl_data = investment.get('implementation', {})

        # Normalize list format → dict format (agent template used lists historically)
        if isinstance(license_data, list):
            license_data = {f'year_{i+1}': v for i, v in enumerate(license_data)}
        if isinstance(impl_data, list):
            impl_data = {f'year_{i+1}': v for i, v in enumerate(impl_data)}

        # Fallback: read from investment_schedule if license/implementation not provided
        if not any(license_data.get(f'year_{i+1}', 0) for i in range(5)) and \
           not any(impl_data.get(f'year_{i+1}', 0) for i in range(5)):
            sched = self.config.get('investment_schedule', {})
            if sched:
                print("⚠ investment.license/implementation not found — deriving from investment_schedule "
                      "(80% implementation / 20% license split). Update config with exact breakdown.")
                # investment_schedule uses year_0 (pre-go-live) through year_4.
                # Map to year_1..5 in the Excel (year_0 cost is rolled into Year 1).
                yr0 = sched.get('year_0', 0)
                for i in range(5):
                    # Excel Year 1 = pre-go-live (year_0) + first operational year (year_1)
                    # Excel Year N (N>1) = schedule year_N
                    combined = (yr0 + sched.get('year_1', 0)) if i == 0 else sched.get(f'year_{i+1}', 0)
                    impl_data[f'year_{i+1}'] = combined * 0.8
                    license_data[f'year_{i+1}'] = combined * 0.2
            else:
                print("⚠ No investment data found (investment.license/implementation or investment_schedule). "
                      "Investment will show as $0. Update the config before presenting to a client.")

        ws[f'B{row}'] = "License"
        ws[f'B{row}'].font = Font(bold=True)
        row += 1
        for yr in range(5):
            ws.cell(row=row, column=2, value=f"  Year {yr + 1}")
            val = license_data.get(f'year_{yr+1}', 0)
            ws.cell(row=row, column=3, value=val)
            ws.cell(row=row, column=3).number_format = '$#,##0'
            ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor=self.COLORS['editable'])
            self.cell_map['model_inputs']['investment'][f'license_y{yr+1}'] = f'C{row}'
            row += 1

        row += 1
        ws[f'B{row}'] = "Implementation"
        ws[f'B{row}'].font = Font(bold=True)
        row += 1
        for yr in range(5):
            ws.cell(row=row, column=2, value=f"  Year {yr + 1}")
            val = impl_data.get(f'year_{yr+1}', 0)
            ws.cell(row=row, column=3, value=val)
            ws.cell(row=row, column=3).number_format = '$#,##0'
            ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor=self.COLORS['editable'])
            self.cell_map['model_inputs']['investment'][f'impl_y{yr+1}'] = f'C{row}'
            row += 1

        # Active Curves (INDEX from Scenario Data)
        row += 2
        ws[f'B{row}'] = "ACTIVE CURVES (auto-selected from scenario)"
        ws[f'B{row}'].font = Font(bold=True, size=12, color=self.COLORS['primary'])
        row += 1
        ws[f'B{row}'] = "These cells use INDEX formulas — they update when you change the scenario on the Cashflows sheet."
        ws[f'B{row}'].font = Font(italic=True, size=9)
        row += 1

        sd_curves = self.cell_map['scenario_data']['curves']
        for cat in sorted(sd_curves.keys()):
            ws[f'B{row}'] = f"{cat.replace('_', ' ').title()} — Implementation"
            ws[f'B{row}'].font = Font(bold=True)
            self.cell_map['model_inputs']['curves'].setdefault(cat, {})
            row += 1
            impl_start = sd_curves[cat]['impl_start']
            for yr in range(5):
                ws.cell(row=row, column=2, value=f"  Year {yr + 1}")
                sd_row = impl_start + yr
                formula = f"=INDEX('Scenario Data'!$C${sd_row}:$E${sd_row},1,{self.sc_cell})"
                ws.cell(row=row, column=3, value=formula)
                ws.cell(row=row, column=3).number_format = '0%'
                self.cell_map['model_inputs']['curves'][cat][f'impl_y{yr+1}'] = f'C{row}'
                row += 1

            ws[f'B{row}'] = f"{cat.replace('_', ' ').title()} — Effectiveness"
            ws[f'B{row}'].font = Font(bold=True)
            row += 1
            eff_start = sd_curves[cat]['eff_start']
            for yr in range(5):
                ws.cell(row=row, column=2, value=f"  Year {yr + 1}")
                sd_row = eff_start + yr
                formula = f"=INDEX('Scenario Data'!$C${sd_row}:$E${sd_row},1,{self.sc_cell})"
                ws.cell(row=row, column=3, value=formula)
                ws.cell(row=row, column=3).number_format = '0%'
                self.cell_map['model_inputs']['curves'][cat][f'eff_y{yr+1}'] = f'C{row}'
                row += 1
            row += 1

    def _find_impact_ref(self, inp_data, driver, group_key):
        """Return a formula that resolves to the scenario-switched backbase_impact value.

        Priority:
        1. Explicit 'impact_key' in inp_data — direct, unambiguous mapping.
        2. model_inputs.impacts — already-written Model Inputs INDEX cell (only populated
           if Active Backbase Impacts was built before this driver, which is not the default
           order; kept as a fallback for future restructuring).
        3. scenario_data.impacts — build an INDEX formula directly over the Scenario Data
           sheet, which is always available since that sheet is written before Model Inputs.
        4. Value-match fallback — matches by moderate-scenario impact value (unreliable when
           two drivers share the same value; kept for backward compat).

        Returns a cell reference string like 'C42' (Model Inputs) or an Excel formula string
        like "INDEX('Scenario Data'!$C$15:$E$15,1,'Cashflows'!$C$5)".
        Callers must prefix with '=' and distinguish between the two using startswith('INDEX').
        """
        val = inp_data.get('value', 0)
        mod_impacts = self.config.get('scenarios', {}).get('moderate', {}).get('backbase_impacts', {})
        sd_impacts = self.cell_map.get('scenario_data', {}).get('impacts', {})
        mi_impacts = self.cell_map.get('model_inputs', {}).get('impacts', {})

        # 1. Explicit impact_key field in the config
        explicit_key = inp_data.get('impact_key')
        if explicit_key:
            if explicit_key in mi_impacts:
                return mi_impacts[explicit_key]
            if explicit_key in sd_impacts:
                sd_row = sd_impacts[explicit_key]
                return f"INDEX('Scenario Data'!$C${sd_row}:$E${sd_row},1,{self.sc_cell})"

        # 2. Already-built Model Inputs cell (populated if section order is changed in future)
        for imp_key, imp_val in mod_impacts.items():
            if abs(imp_val - val) < 0.001 and imp_key in mi_impacts:
                return mi_impacts[imp_key]

        # 3. Build INDEX formula directly over Scenario Data (always available)
        for imp_key, imp_val in mod_impacts.items():
            if abs(imp_val - val) < 0.001 and imp_key in sd_impacts:
                sd_row = sd_impacts[imp_key]
                return f"INDEX('Scenario Data'!$C${sd_row}:$E${sd_row},1,{self.sc_cell})"

        return None

    # ── Journey Analysis Sheet (all lever groups in one sheet) ───────

    def _create_journey_analysis_sheet(self):
        """Create Journey Analysis sheet showing full formula chain per driver:
        Input factors (from Model Inputs) → Baseline formula → Backbase Impact → Annual Benefit.
        """
        ws = self.wb.create_sheet("Journey Analysis", self.sheet_index)
        self.sheet_index += 1
        self.lever_sheet_names = {}

        for col, w in [('A', 5), ('B', 45), ('C', 20), ('D', 25)]:
            ws.column_dimensions[col].width = w

        ws['B1'] = "Journey Analysis — Benefit Calculation Chain"
        ws['B1'].font = Font(bold=True, size=18, color=self.COLORS['primary'])

        # Active Scenario — read-only reference to Cashflows dropdown
        ws['B2'] = "Active Scenario:"
        ws['B2'].font = Font(bold=True, size=11, color=self.COLORS['primary'])
        sc_label_formula = (
            f"=IF({self.sc_cell}=1,\"Conservative\","
            f"IF({self.sc_cell}=2,\"Moderate\",\"Aggressive\"))"
        )
        ws['C2'] = sc_label_formula
        ws['C2'].font = Font(bold=True, size=11, color=self.COLORS['positive'])
        ws['D2'] = "(Change on Cashflows sheet)"
        ws['D2'].font = Font(italic=True, size=9)

        ws['B3'] = ("Each lever: input factors → Baseline (formula) → × Backbase Impact → Annual Benefit. "
                    "Change scenario on Cashflows to update scenario-linked impacts.")
        ws['B3'].font = Font(italic=True, size=10)

        self.cell_map.setdefault('lever_sheets', {})
        mi_groups = self.cell_map.get('model_inputs', {}).get('groups', {})
        mi_impacts = self.cell_map.get('model_inputs', {}).get('impacts', {})

        row = 5
        group_total_rows = []

        for group_key in self.group_order:
            group = self.lever_groups[group_key]
            self.cell_map['lever_sheets'][group_key] = {}

            # Group header
            ws[f'B{row}'] = group.get('group_name', group_key)
            ws[f'B{row}'].font = Font(bold=True, size=13, color=self.COLORS['primary'])
            row += 1
            evidence = group.get('evidence_ids', [])
            if evidence:
                ws[f'B{row}'] = f"Evidence: {', '.join(evidence)}"
                ws[f'B{row}'].font = Font(italic=True, size=9)
                row += 1

            benefit_cells = []

            def _write_driver_chain(drv_key, driver, driver_type_label):
                """Write the full formula chain for one driver. Returns the annual benefit cell ref."""
                nonlocal row

                # Driver sub-header
                ws.cell(row=row, column=2, value=f"  {driver.get('name', drv_key)}")
                ws.cell(row=row, column=2).font = Font(bold=True, size=10,
                                                       color=self.COLORS['positive'] if driver_type_label == 'Revenue'
                                                       else self.COLORS['primary'])
                ws.cell(row=row, column=4, value=driver_type_label)
                ws.cell(row=row, column=4).font = Font(italic=True, size=9)
                row += 1

                # Column headers for this driver's factor block
                self._write_header_row(ws, row, ['Factor', 'Value', 'Unit'])
                row += 1

                # Input factor rows — reference Model Inputs cells
                # Skip 'backbase_impact' here; it gets its own dedicated row below
                mi_driver_inputs = mi_groups.get(group_key, {}).get('inputs', {})
                ja_driver_inputs = {}
                for inp_key, inp_data in driver.get('inputs', {}).items():
                    if inp_key == 'backbase_impact':
                        continue  # handled separately in the Backbase Impact row
                    mi_cell = mi_driver_inputs.get(f'{drv_key}__{inp_key}')
                    ws.cell(row=row, column=2, value=f"    {inp_key.replace('_', ' ').title()}")
                    if mi_cell:
                        ws.cell(row=row, column=3, value=f"='Model Inputs'!{mi_cell}")
                    else:
                        ws.cell(row=row, column=3, value=inp_data.get('value', 0))
                    # Format numbers: %, currency, or plain
                    val = inp_data.get('value', 0)
                    unit = inp_data.get('unit', '')
                    if isinstance(val, float) and val < 1 and 'percentage' in unit.lower():
                        ws.cell(row=row, column=3).number_format = '0.0%'
                    elif isinstance(val, (int, float)) and abs(val) >= 1000:
                        ws.cell(row=row, column=3).number_format = '#,##0'
                    ws.cell(row=row, column=4, value=inp_data.get('unit', ''))
                    ws.cell(row=row, column=4).font = Font(size=9)
                    ja_driver_inputs[inp_key] = f'C{row}'
                    row += 1

                # Baseline row — formula if template exists
                ws.cell(row=row, column=2, value="    Baseline")
                ws.cell(row=row, column=2).font = Font(bold=True)
                formula_tpl = driver.get('baseline_formula', '')
                if formula_tpl:
                    excel_formula = '=' + formula_tpl
                    for key, cell_ref in ja_driver_inputs.items():
                        excel_formula = excel_formula.replace(f'{{{key}}}', cell_ref)
                    ws.cell(row=row, column=3, value=excel_formula)
                else:
                    # Fall back to referencing Model Inputs baseline cell
                    mi_baseline = mi_groups.get(group_key, {}).get('drivers', {}).get(f'{drv_key}__baseline')
                    if mi_baseline:
                        ws.cell(row=row, column=3, value=f"='Model Inputs'!{mi_baseline}")
                    else:
                        ws.cell(row=row, column=3, value=driver.get('baseline_annual', 0))
                ws.cell(row=row, column=3).number_format = '#,##0'
                ws.cell(row=row, column=3).font = Font(bold=True)
                ws.cell(row=row, column=4, value=self.config.get('currency', 'NGN') + '/year')
                ws.cell(row=row, column=4).font = Font(size=9)
                baseline_row = row
                row += 1

                # Backbase Impact row — look up the Model Inputs cell directly
                ws.cell(row=row, column=2, value="    Backbase Impact")
                ws.cell(row=row, column=2).font = Font(bold=True)
                mi_bi_cell = mi_driver_inputs.get(f'{drv_key}__backbase_impact')
                if mi_bi_cell:
                    # Check if this cell is an INDEX formula (scenario-linked via Active Backbase Impacts)
                    if mi_bi_cell in mi_impacts.values():
                        ws.cell(row=row, column=3, value=f"='Model Inputs'!{mi_bi_cell}")
                    else:
                        ws.cell(row=row, column=3, value=f"='Model Inputs'!{mi_bi_cell}")
                else:
                    ws.cell(row=row, column=3, value=0.30)
                    ws.cell(row=row, column=5, value="⚠ FALLBACK — scenario linkage missing, defaulted to 30%")
                    ws.cell(row=row, column=5).font = Font(size=9, color='FF4444', italic=True)
                ws.cell(row=row, column=3).number_format = '0.0%'
                ws.cell(row=row, column=3).font = Font(bold=True)
                ws.cell(row=row, column=4, value="% of baseline captured by Backbase")
                ws.cell(row=row, column=4).font = Font(size=9, italic=True)
                backbase_row = row
                row += 1

                # Annual Benefit row
                ws.cell(row=row, column=2, value="    Annual Benefit")
                ws.cell(row=row, column=2).font = Font(bold=True, size=11)
                ws.cell(row=row, column=3, value=f"=C{baseline_row}*C{backbase_row}")
                ws.cell(row=row, column=3).number_format = '#,##0'
                ws.cell(row=row, column=3).font = Font(bold=True, size=11, color=self.COLORS['positive'])
                ws.cell(row=row, column=4, value=self.config.get('currency', 'NGN') + '/year (scenario-adjusted)')
                ws.cell(row=row, column=4).font = Font(size=9, color=self.COLORS['positive'])
                benefit_cell = f'C{row}'
                row += 2  # blank line between drivers
                return benefit_cell

            # Revenue Drivers
            rev_drivers = group.get('revenue_drivers', {})
            if rev_drivers:
                ws[f'B{row}'] = "Revenue Generation"
                ws[f'B{row}'].font = Font(bold=True, size=11, color=self.COLORS['positive'])
                row += 1
                for drv_key, driver in rev_drivers.items():
                    cell = _write_driver_chain(drv_key, driver, 'Revenue')
                    benefit_cells.append(cell)

            # Cost Drivers
            cost_drivers = group.get('cost_drivers', {})
            if cost_drivers:
                ws[f'B{row}'] = "Cost Reduction"
                ws[f'B{row}'].font = Font(bold=True, size=11, color=self.COLORS['primary'])
                row += 1
                for drv_key, driver in cost_drivers.items():
                    cell = _write_driver_chain(drv_key, driver, 'Cost')
                    benefit_cells.append(cell)

            # Servicing reference (if this group has servicing_analysis)
            servicing = group.get('servicing_analysis')
            if servicing:
                ws[f'B{row}'] = "Servicing Cost Avoidance"
                ws[f'B{row}'].font = Font(bold=True, size=11, color=self.COLORS['primary'])
                row += 1
                ws.cell(row=row, column=2, value="  Servicing Total (see Servicing Detail sheet)")
                svc_ref = self.cell_map.get('servicing_grand_total')
                if svc_ref:
                    ws.cell(row=row, column=3, value=f"='Servicing Detail'!{svc_ref}")
                else:
                    ws.cell(row=row, column=3, value=group.get('totals', {}).get('total_cost_saved', 0))
                ws.cell(row=row, column=3).number_format = '#,##0'
                benefit_cells.append(f'C{row}')
                row += 1

            # Group total
            row += 1
            ws.cell(row=row, column=2, value=f"Total: {group.get('group_name', group_key)}")
            ws.cell(row=row, column=2).font = Font(bold=True, size=11)
            if benefit_cells:
                ws.cell(row=row, column=3, value=f"={'+'.join(benefit_cells)}")
            ws.cell(row=row, column=3).number_format = '#,##0'
            ws.cell(row=row, column=3).font = Font(bold=True, size=11, color=self.COLORS['positive'])

            self.cell_map['lever_sheets'][group_key]['annual_total'] = f'C{row}'
            group_total_rows.append(f'C{row}')
            row += 3

        # Grand Total
        ws[f'B{row}'] = "GRAND TOTAL — ALL LEVERS"
        ws[f'B{row}'].font = Font(bold=True, size=14)
        row += 1
        self._write_header_row(ws, row, ['', 'Annual Benefit', ''])
        row += 1
        ws.cell(row=row, column=2, value="Total Annual Benefit (Scenario-Adjusted, Steady State)")
        ws.cell(row=row, column=2).font = Font(bold=True, size=12)
        if group_total_rows:
            ws.cell(row=row, column=3, value=f"={'+'.join(group_total_rows)}")
        ws.cell(row=row, column=3).number_format = '#,##0'
        ws.cell(row=row, column=3).font = Font(bold=True, size=14, color=self.COLORS['positive'])
        self.cell_map['journey_analysis_grand_total'] = f'C{row}'

    # ── Servicing Detail Sheet ────────────────────────────────────────

    def _create_servicing_detail_sheet(self):
        ws = self.wb.create_sheet("Servicing Detail", self.sheet_index)
        self.sheet_index += 1

        for col, w in [('A', 5), ('B', 30), ('C', 15), ('D', 15), ('E', 12), ('F', 18), ('G', 14), ('H', 14), ('I', 16), ('J', 16), ('K', 16), ('L', 40)]:
            ws.column_dimensions[col].width = w

        ws['B1'] = "Customer Servicing — Channel Analysis"
        ws['B1'].font = Font(bold=True, size=18, color=self.COLORS['primary'])

        ws['B3'] = "Vol Saving = Baseline × Vol Deflection. Time Saving = Remaining Vol × Time × Time Reduction × Rate. Total = Vol + Time Saving."
        ws['B3'].font = Font(italic=True, size=10)

        row = 5
        self._write_header_row(ws, row, ['Task', 'Volume', 'Time (hrs)', 'Rate', 'Baseline', 'Vol Deflection', 'Time Reduction', 'Vol Saving', 'Time Saving', 'Total Saved', 'Source'])
        row += 1

        # Find the servicing group
        for group_key, group in self.lever_groups.items():
            servicing = group.get('servicing_analysis')
            if not servicing:
                continue

            evidence = group.get('evidence_ids', [])
            ws.cell(row=row, column=2, value=f"Evidence: {', '.join(evidence)}")
            ws.cell(row=row, column=2).font = Font(italic=True)
            row += 1

            channel_subtotal_cells = []
            for channel_key in ['branch', 'call_center', 'back_office']:
                channel = servicing.get(channel_key, {})
                if not channel:
                    continue

                row += 1
                ws[f'B{row}'] = channel.get('channel_name', channel_key.title())
                ws[f'B{row}'].font = Font(bold=True, size=12, color=self.COLORS['primary'])
                row += 1

                task_saved_cells = []
                task_baseline_cells = []
                task_volume_cells = []

                for task in channel.get('tasks', []):
                    ws.cell(row=row, column=2, value=task.get('task', ''))
                    # Volume (editable)
                    ws.cell(row=row, column=3, value=task.get('yearly_volume', 0))
                    ws.cell(row=row, column=3).number_format = '#,##0'
                    ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor=self.COLORS['editable'])
                    # Time (editable)
                    ws.cell(row=row, column=4, value=task.get('time_spent_hours', 0))
                    ws.cell(row=row, column=4).number_format = '0.00'
                    ws.cell(row=row, column=4).fill = PatternFill("solid", fgColor=self.COLORS['editable'])
                    # Rate (editable)
                    ws.cell(row=row, column=5, value=task.get('fte_rate', 25))
                    ws.cell(row=row, column=5).number_format = '$#,##0'
                    ws.cell(row=row, column=5).fill = PatternFill("solid", fgColor=self.COLORS['editable'])
                    # Baseline = Volume × Time × Rate (FORMULA)
                    ws.cell(row=row, column=6, value=f"=C{row}*D{row}*E{row}")
                    ws.cell(row=row, column=6).number_format = '$#,##0'
                    task_baseline_cells.append(f'F{row}')
                    task_volume_cells.append(f'C{row}')
                    # Volume Deflection Rate (editable)
                    vdr = task.get('volume_deflection_rate', task.get('backbase_impact', 0))
                    ws.cell(row=row, column=7, value=vdr)
                    ws.cell(row=row, column=7).number_format = '0%'
                    ws.cell(row=row, column=7).fill = PatternFill("solid", fgColor=self.COLORS['editable'])
                    source = task.get('impact_source', '')
                    if 'BACKBASE' in source.upper():
                        ws.cell(row=row, column=7).fill = PatternFill("solid", fgColor=self.COLORS['benchmark'])
                    # Time Reduction Rate (editable)
                    trr = task.get('time_reduction_rate', 0)
                    ws.cell(row=row, column=8, value=trr)
                    ws.cell(row=row, column=8).number_format = '0%'
                    ws.cell(row=row, column=8).fill = PatternFill("solid", fgColor=self.COLORS['editable'])
                    if 'BACKBASE' in source.upper():
                        ws.cell(row=row, column=8).fill = PatternFill("solid", fgColor=self.COLORS['benchmark'])
                    # Vol Saving = Baseline × Vol Deflection (FORMULA)
                    ws.cell(row=row, column=9, value=f"=F{row}*G{row}")
                    ws.cell(row=row, column=9).number_format = '$#,##0'
                    ws.cell(row=row, column=9).font = Font(color=self.COLORS['positive'])
                    # Time Saving = Volume × (1-VolDeflection) × Time × TimeReduction × Rate (FORMULA)
                    ws.cell(row=row, column=10, value=f"=C{row}*(1-G{row})*D{row}*H{row}*E{row}")
                    ws.cell(row=row, column=10).number_format = '$#,##0'
                    ws.cell(row=row, column=10).font = Font(color=self.COLORS['positive'])
                    # Total Saved = Vol Saving + Time Saving (FORMULA)
                    ws.cell(row=row, column=11, value=f"=I{row}+J{row}")
                    ws.cell(row=row, column=11).number_format = '$#,##0'
                    ws.cell(row=row, column=11).font = Font(bold=True, color=self.COLORS['positive'])
                    task_saved_cells.append(f'K{row}')
                    # Source
                    ws.cell(row=row, column=12, value=task.get('volume_source', ''))
                    ws.cell(row=row, column=12).font = Font(size=8, color='666666')
                    row += 1

                # Channel subtotal (FORMULA)
                ws.cell(row=row, column=2, value=f"{channel.get('channel_name', '')} Subtotal")
                ws.cell(row=row, column=2).font = Font(bold=True)
                if task_volume_cells:
                    ws.cell(row=row, column=3, value=f"=SUM({task_volume_cells[0]}:{task_volume_cells[-1]})")
                ws.cell(row=row, column=3).number_format = '#,##0'
                ws.cell(row=row, column=3).font = Font(bold=True)
                if task_baseline_cells:
                    ws.cell(row=row, column=6, value=f"=SUM({task_baseline_cells[0]}:{task_baseline_cells[-1]})")
                ws.cell(row=row, column=6).number_format = '$#,##0'
                ws.cell(row=row, column=6).font = Font(bold=True)
                if task_saved_cells:
                    # Vol Saving subtotal
                    vol_cells = [c.replace('K', 'I') for c in task_saved_cells]
                    ws.cell(row=row, column=9, value=f"=SUM({vol_cells[0]}:{vol_cells[-1]})")
                    ws.cell(row=row, column=9).number_format = '$#,##0'
                    ws.cell(row=row, column=9).font = Font(bold=True)
                    # Time Saving subtotal
                    time_cells = [c.replace('K', 'J') for c in task_saved_cells]
                    ws.cell(row=row, column=10, value=f"=SUM({time_cells[0]}:{time_cells[-1]})")
                    ws.cell(row=row, column=10).number_format = '$#,##0'
                    ws.cell(row=row, column=10).font = Font(bold=True)
                    # Total Saved subtotal
                    ws.cell(row=row, column=11, value=f"=SUM({task_saved_cells[0]}:{task_saved_cells[-1]})")
                    ws.cell(row=row, column=11).number_format = '$#,##0'
                    ws.cell(row=row, column=11).font = Font(bold=True, color=self.COLORS['positive'])
                channel_subtotal_cells.append(f'K{row}')
                row += 1

            # Grand total across all channels
            if channel_subtotal_cells:
                row += 1
                ws.cell(row=row, column=2, value="TOTAL SERVICING SAVINGS")
                ws.cell(row=row, column=2).font = Font(bold=True, size=12)
                ws.cell(row=row, column=6, value=f"=SUM({','.join(c.replace('K','F') for c in channel_subtotal_cells)})")
                ws.cell(row=row, column=6).number_format = '$#,##0'
                ws.cell(row=row, column=6).font = Font(bold=True, size=12)
                # Vol Saving total
                ws.cell(row=row, column=9, value=f"={'+'.join(c.replace('K','I') for c in channel_subtotal_cells)}")
                ws.cell(row=row, column=9).number_format = '$#,##0'
                ws.cell(row=row, column=9).font = Font(bold=True, size=12)
                # Time Saving total
                ws.cell(row=row, column=10, value=f"={'+'.join(c.replace('K','J') for c in channel_subtotal_cells)}")
                ws.cell(row=row, column=10).number_format = '$#,##0'
                ws.cell(row=row, column=10).font = Font(bold=True, size=12)
                # Total Saved
                ws.cell(row=row, column=11, value=f"={'+'.join(channel_subtotal_cells)}")
                ws.cell(row=row, column=11).number_format = '$#,##0'
                ws.cell(row=row, column=11).font = Font(bold=True, size=12, color=self.COLORS['positive'])
                self.cell_map['servicing_grand_total'] = f'K{row}'
                grand_total_row = row

            # Growth-adjusted cost avoidance section (if yoy_growth_rate in config)
            yoy_rate = self.config.get('yoy_growth_rate', 0)
            if not yoy_rate:
                # Also check backbase_loading.yoy_growth (array format)
                yoy_arr = self.config.get('backbase_loading', {}).get('yoy_growth', [])
                if yoy_arr:
                    yoy_rate = yoy_arr[0] if isinstance(yoy_arr, list) else yoy_arr
            if yoy_rate and yoy_rate > 0 and channel_subtotal_cells:
                annual_hrs = self.config.get('annual_hours_per_fte', 1800)
                loaded_fte = self.config.get('loaded_annual_fte_cost', 55000)
                vol_total_ref = grand_total_row  # row with total volume

                row += 3
                ws[f'B{row}'] = "GROWTH-ADJUSTED COST AVOIDANCE"
                ws[f'B{row}'].font = Font(bold=True, size=14, color=self.COLORS['primary'])
                row += 1
                ws[f'B{row}'] = f"Projecting volume growth at {yoy_rate:.0%}/year — staff NOT needed due to digital handling"
                ws[f'B{row}'].font = Font(italic=True, size=10)
                row += 2

                # Parameters row
                ws.cell(row=row, column=2, value="YoY Growth Rate")
                ws.cell(row=row, column=3, value=yoy_rate)
                ws.cell(row=row, column=3).number_format = '0%'
                ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor=self.COLORS['editable'])
                growth_rate_cell = f'C{row}'
                row += 1
                ws.cell(row=row, column=2, value="Annual Hrs per FTE")
                ws.cell(row=row, column=3, value=annual_hrs)
                ws.cell(row=row, column=3).number_format = '#,##0'
                ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor=self.COLORS['editable'])
                hrs_cell = f'C{row}'
                row += 1
                ws.cell(row=row, column=2, value="Loaded Annual FTE Cost")
                ws.cell(row=row, column=3, value=loaded_fte)
                ws.cell(row=row, column=3).number_format = '$#,##0'
                ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor=self.COLORS['editable'])
                fte_cost_cell = f'C{row}'
                row += 1
                # Weighted avg time per interaction
                ws.cell(row=row, column=2, value="Weighted Avg Time/Interaction (hrs)")
                # Compute from config: sum(vol*time) / sum(vol) across all tasks
                total_vol = 0
                total_vol_time = 0
                for gk, grp in self.lever_groups.items():
                    sa = grp.get('servicing_analysis')
                    if not sa:
                        continue
                    for ch_key in ['branch', 'call_center', 'back_office']:
                        ch = sa.get(ch_key, {})
                        for t in ch.get('tasks', []):
                            v = t.get('yearly_volume', 0)
                            h = t.get('time_spent_hours', 0)
                            total_vol += v
                            total_vol_time += v * h
                avg_time = total_vol_time / total_vol if total_vol > 0 else 0.25
                ws.cell(row=row, column=3, value=round(avg_time, 3))
                ws.cell(row=row, column=3).number_format = '0.000'
                avg_time_cell = f'C{row}'
                row += 1
                # Weighted avg combined impact (from dual dimensions)
                ws.cell(row=row, column=2, value="Weighted Avg Combined Impact")
                total_vol2 = 0
                total_combined = 0
                for gk, grp in self.lever_groups.items():
                    sa = grp.get('servicing_analysis')
                    if not sa:
                        continue
                    for ch_key in ['branch', 'call_center', 'back_office']:
                        ch = sa.get(ch_key, {})
                        for t in ch.get('tasks', []):
                            v = t.get('yearly_volume', 0)
                            vdr = t.get('volume_deflection_rate', t.get('backbase_impact', 0))
                            trr = t.get('time_reduction_rate', 0)
                            combined = 1 - (1 - vdr) * (1 - trr)
                            total_vol2 += v
                            total_combined += v * combined
                avg_impact = total_combined / total_vol2 if total_vol2 > 0 else 0.3
                ws.cell(row=row, column=3, value=round(avg_impact, 3))
                ws.cell(row=row, column=3).number_format = '0.0%'
                avg_impact_cell = f'C{row}'
                row += 2

                # Year-by-year growth avoidance table
                self._write_header_row(ws, row, ['Metric', 'Y1', 'Y2', 'Y3', 'Y4', 'Y5'])
                row += 1
                base_vol_row = row
                ws.cell(row=row, column=2, value="Base Volume (current)")
                # Use grand total volume from the subtotals
                base_vol_cells = [c.replace('K', 'C') for c in channel_subtotal_cells]
                ws.cell(row=row, column=3, value=f"={'+'.join(base_vol_cells)}")
                ws.cell(row=row, column=3).number_format = '#,##0'
                base_vol_ref = f'C{row}'
                for yr in range(1, 5):
                    ws.cell(row=row, column=3 + yr, value=f"={base_vol_ref}")
                    ws.cell(row=row, column=3 + yr).number_format = '#,##0'
                row += 1

                growth_vol_row = row
                ws.cell(row=row, column=2, value="Growth Volume (incremental)")
                for yr in range(5):
                    col = 3 + yr
                    ws.cell(row=row, column=col, value=f"={base_vol_ref}*((1+{growth_rate_cell})^{yr+1}-1)")
                    ws.cell(row=row, column=col).number_format = '#,##0'
                row += 1

                digital_row = row
                ws.cell(row=row, column=2, value="Handled Digitally")
                for yr in range(5):
                    col = 3 + yr
                    ws.cell(row=row, column=col, value=f"={get_column_letter(col)}{growth_vol_row}*{avg_impact_cell}")
                    ws.cell(row=row, column=col).number_format = '#,##0'
                row += 1

                fte_row = row
                ws.cell(row=row, column=2, value="FTEs Avoided")
                for yr in range(5):
                    col = 3 + yr
                    ws.cell(row=row, column=col, value=f"={get_column_letter(col)}{digital_row}*{avg_time_cell}/{hrs_cell}")
                    ws.cell(row=row, column=col).number_format = '0.0'
                    ws.cell(row=row, column=col).font = Font(bold=True, color=self.COLORS['positive'])
                row += 1

                cost_avoided_row = row
                ws.cell(row=row, column=2, value="Cost Avoided (growth)")
                ws.cell(row=row, column=2).font = Font(bold=True)
                for yr in range(5):
                    col = 3 + yr
                    ws.cell(row=row, column=col, value=f"={get_column_letter(col)}{fte_row}*{fte_cost_cell}")
                    ws.cell(row=row, column=col).number_format = '$#,##0'
                    ws.cell(row=row, column=col).font = Font(bold=True, color=self.COLORS['positive'])
                self.cell_map['servicing_growth_avoidance_row'] = cost_avoided_row

    # ── Cashflows Sheet (merged with Results Dashboard) ──────────────

    def _create_cashflows_sheet(self):
        """Single Cashflows sheet: financial summary at top, year-by-year projections,
        investment, net cashflow, financial metrics, and benefits-by-lever breakdown."""
        ws = self.wb.create_sheet("Cashflows", self.sheet_index)
        self.sheet_index += 1

        for col, w in [('A', 5), ('B', 45)] + [(get_column_letter(i), 18) for i in range(3, 9)]:
            ws.column_dimensions[col].width = w

        ws['B1'] = "5 Year Cashflows & ROI"
        ws['B1'].font = Font(bold=True, size=18, color=self.COLORS['primary'])

        ws['B3'] = "All values are live formulas. Change the scenario below — Journey Analysis and Cashflows update automatically."
        ws['B3'].font = Font(italic=True, size=10)

        # ── Scenario Selector ──
        # IMPORTANT: self.sc_cell is pre-declared as "'Cashflows'!$C$5" in generate().
        # The dropdown MUST be placed at row 5, column 3 to match that reference.
        row = 4
        ws[f'B{row}'] = "SCENARIO SELECTOR"
        ws[f'B{row}'].font = Font(bold=True, size=12, color=self.COLORS['primary'])
        row += 1  # row = 5
        ws[f'B{row}'] = "Active Scenario (1=Conservative, 2=Moderate, 3=Aggressive)"
        sc_map = {'conservative': 1, 'moderate': 2, 'aggressive': 3}
        default_sc = sc_map.get(self.config.get('selected_scenario', 'Moderate').lower(), 2)
        ws.cell(row=row, column=3, value=default_sc)
        ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor=self.COLORS['editable'])
        ws.cell(row=row, column=3).font = Font(bold=True, size=14)
        dv = DataValidation(type="list", formula1='"1,2,3"', allow_blank=False)
        dv.error = "Select 1 (Conservative), 2 (Moderate), or 3 (Aggressive)"
        dv.errorTitle = "Invalid Scenario"
        ws.add_data_validation(dv)
        dv.add(ws.cell(row=row, column=3))
        self.cell_map.setdefault('cashflows', {})
        self.cell_map['cashflows']['scenario_row'] = row  # row = 5, matching self.sc_cell = "'Cashflows'!$C$5"
        row += 1  # row = 6
        ws[f'B{row}'] = (
            f"=IF(C5=1,\"Conservative — slower adoption, cautious assumptions\","
            f"IF(C5=2,\"Moderate — base case, phased rollout\",\"Aggressive — fast adoption, full benefit\"))"
        )
        ws[f'B{row}'].font = Font(italic=True, size=10, color=self.COLORS['positive'])

        row += 2  # blank row, then year headers at row 8

        # ── Year-by-year headers ──
        self._write_header_row(ws, row, ['', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5', 'Total'])
        row += 1

        # ── Cash Inflows ── (row continues from year headers above)
        ws[f'B{row}'] = "CASH INFLOWS (Benefits)"
        ws[f'B{row}'].font = Font(bold=True, color=self.COLORS['positive'])
        row += 1

        inflow_first_row = row
        for group_key in self.group_order:
            group = self.lever_groups[group_key]
            lever_cells = self.cell_map.get('lever_sheets', {}).get(group_key, {})
            annual_cell = lever_cells.get('annual_total')

            ws.cell(row=row, column=2, value=group.get('group_name', group_key))

            # Year-by-year: annual_benefit × impl_curve × eff_curve
            curve_cat = self._get_curve_category(group_key)
            curve_map = self.cell_map['model_inputs'].get('curves', {}).get(curve_cat, {})

            for yr in range(5):
                impl_ref = curve_map.get(f'impl_y{yr+1}')
                eff_ref = curve_map.get(f'eff_y{yr+1}')
                if annual_cell and impl_ref and eff_ref:
                    formula = f"='Journey Analysis'!{annual_cell}*'Model Inputs'!{impl_ref}*'Model Inputs'!{eff_ref}"
                    ws.cell(row=row, column=3 + yr, value=formula)
                else:
                    ws.cell(row=row, column=3 + yr, value=0)
                ws.cell(row=row, column=3 + yr).number_format = '$#,##0'

            ws.cell(row=row, column=8, value=f"=SUM(C{row}:G{row})")
            ws.cell(row=row, column=8).number_format = '$#,##0'
            row += 1

        # Growth avoidance inflow row if it exists
        growth_row = self.cell_map.get('servicing_growth_avoidance_row')
        if growth_row:
            ws.cell(row=row, column=2, value="Servicing Growth Avoidance (FTE)")
            for yr in range(5):
                svc_col = get_column_letter(3 + yr)
                ws.cell(row=row, column=3 + yr, value=f"='Servicing Detail'!{svc_col}{growth_row}")
                ws.cell(row=row, column=3 + yr).number_format = '$#,##0'
            ws.cell(row=row, column=8, value=f"=SUM(C{row}:G{row})")
            ws.cell(row=row, column=8).number_format = '$#,##0'
            row += 1

        inflow_last_row = row - 1

        # Total Inflows
        row += 1
        ws.cell(row=row, column=2, value="Total Cash Inflow")
        ws.cell(row=row, column=2).font = Font(bold=True)
        inflow_total_row = row
        for col_idx in range(3, 9):
            col_letter = get_column_letter(col_idx)
            ws.cell(row=row, column=col_idx, value=f"=SUM({col_letter}{inflow_first_row}:{col_letter}{inflow_last_row})")
            ws.cell(row=row, column=col_idx).number_format = '$#,##0'
            ws.cell(row=row, column=col_idx).font = Font(bold=True)

        # ── Cash Outflows ──
        row += 3
        ws[f'B{row}'] = "CASH OUTFLOWS (Investment)"
        ws[f'B{row}'].font = Font(bold=True, color=self.COLORS['negative'])
        row += 1

        ws.cell(row=row, column=2, value="License")
        license_row = row
        for yr in range(5):
            lic_ref = self.cell_map['model_inputs']['investment'].get(f'license_y{yr+1}')
            if lic_ref:
                ws.cell(row=row, column=3 + yr, value=f"='Model Inputs'!{lic_ref}")
            ws.cell(row=row, column=3 + yr).number_format = '$#,##0'
        ws.cell(row=row, column=8, value=f"=SUM(C{row}:G{row})")
        ws.cell(row=row, column=8).number_format = '$#,##0'
        row += 1

        ws.cell(row=row, column=2, value="Implementation")
        impl_row = row
        for yr in range(5):
            impl_ref = self.cell_map['model_inputs']['investment'].get(f'impl_y{yr+1}')
            if impl_ref:
                ws.cell(row=row, column=3 + yr, value=f"='Model Inputs'!{impl_ref}")
            ws.cell(row=row, column=3 + yr).number_format = '$#,##0'
        ws.cell(row=row, column=8, value=f"=SUM(C{row}:G{row})")
        ws.cell(row=row, column=8).number_format = '$#,##0'
        row += 1

        ws.cell(row=row, column=2, value="Total Cash Outflow")
        ws.cell(row=row, column=2).font = Font(bold=True)
        outflow_total_row = row
        for col_idx in range(3, 9):
            col_letter = get_column_letter(col_idx)
            ws.cell(row=row, column=col_idx, value=f"={col_letter}{license_row}+{col_letter}{impl_row}")
            ws.cell(row=row, column=col_idx).number_format = '$#,##0'
            ws.cell(row=row, column=col_idx).font = Font(bold=True)

        # ── Net Cashflow ──
        row += 3
        ws[f'B{row}'] = "NET CASHFLOW"
        ws[f'B{row}'].font = Font(bold=True, size=14)
        net_row = row
        for col_idx in range(3, 8):
            col_letter = get_column_letter(col_idx)
            ws.cell(row=row, column=col_idx, value=f"={col_letter}{inflow_total_row}-{col_letter}{outflow_total_row}")
            ws.cell(row=row, column=col_idx).number_format = '$#,##0'
            ws.cell(row=row, column=col_idx).font = Font(bold=True)
        ws.cell(row=row, column=8, value=f"=SUM(C{row}:G{row})")
        ws.cell(row=row, column=8).number_format = '$#,##0'
        ws.cell(row=row, column=8).font = Font(bold=True, size=14)

        # Cumulative Cashflow
        row += 1
        ws.cell(row=row, column=2, value="Cumulative Cashflow")
        cum_row = row
        ws.cell(row=row, column=3, value=f"=C{net_row}")
        ws.cell(row=row, column=3).number_format = '$#,##0'
        for yr in range(1, 5):
            col_idx = 3 + yr
            prev_col = get_column_letter(col_idx - 1)
            cur_col = get_column_letter(col_idx)
            ws.cell(row=row, column=col_idx, value=f"={prev_col}{cum_row}+{cur_col}{net_row}")
            ws.cell(row=row, column=col_idx).number_format = '$#,##0'

        # ── Financial Metrics ──
        row += 4
        ws[f'B{row}'] = "FINANCIAL METRICS"
        ws[f'B{row}'].font = Font(bold=True, size=14, color=self.COLORS['primary'])

        dr_ref = self.cell_map['model_inputs']['discount_rate']

        row += 1
        ws.cell(row=row, column=2, value="Discount Rate (WACC)")
        ws.cell(row=row, column=3, value=f"='Model Inputs'!{dr_ref}")
        ws.cell(row=row, column=3).number_format = '0.0%'
        dr_cell_row = row

        row += 1
        ws.cell(row=row, column=2, value="Total 5-Year Benefits")
        ws.cell(row=row, column=3, value=f"=H{inflow_total_row}")
        ws.cell(row=row, column=3).number_format = '$#,##0'
        ws.cell(row=row, column=3).font = Font(bold=True, size=14)
        benefits_total_row = row

        row += 1
        ws.cell(row=row, column=2, value="Total 5-Year Investment")
        ws.cell(row=row, column=3, value=f"=H{outflow_total_row}")
        ws.cell(row=row, column=3).number_format = '$#,##0'
        invest_total_row = row

        row += 1
        ws.cell(row=row, column=2, value="Net Benefit")
        ws.cell(row=row, column=3, value=f"=C{benefits_total_row}-C{invest_total_row}")
        ws.cell(row=row, column=3).number_format = '$#,##0'
        ws.cell(row=row, column=3).font = Font(bold=True, size=14)

        row += 1
        ws.cell(row=row, column=2, value="ROI %")
        ws.cell(row=row, column=3, value=f"=IFERROR((C{benefits_total_row}-C{invest_total_row})/C{invest_total_row},0)")
        ws.cell(row=row, column=3).number_format = '0.0%'
        ws.cell(row=row, column=3).font = Font(bold=True, size=14, color=self.COLORS['positive'])

        row += 1
        ws.cell(row=row, column=2, value="Net Present Value (NPV)")
        ws.cell(row=row, column=3, value=f"=NPV(C{dr_cell_row},C{net_row}:G{net_row})")
        ws.cell(row=row, column=3).number_format = '$#,##0'
        ws.cell(row=row, column=3).font = Font(bold=True, size=14)

        row += 1
        ws.cell(row=row, column=2, value="Internal Rate of Return (IRR)")
        ws.cell(row=row, column=3, value=f'=IFERROR(IRR(C{net_row}:G{net_row}),"N/A")')
        ws.cell(row=row, column=3).number_format = '0.0%'
        ws.cell(row=row, column=3).font = Font(bold=True, size=14)

        row += 1
        ws.cell(row=row, column=2, value="Payback Period (Years)")
        payback_formula = (
            f"=IF(C{cum_row}>=0,1-((C{cum_row}-C{net_row})/C{net_row}),"
            f"IF(D{cum_row}>=0,2-((D{cum_row}-D{net_row})/D{net_row}),"
            f"IF(E{cum_row}>=0,3-((E{cum_row}-E{net_row})/E{net_row}),"
            f"IF(F{cum_row}>=0,4-((F{cum_row}-F{net_row})/F{net_row}),"
            f"IF(G{cum_row}>=0,5-((G{cum_row}-G{net_row})/G{net_row}),"
            f'"N/A")))))'
        )
        ws.cell(row=row, column=3, value=payback_formula)
        ws.cell(row=row, column=3).number_format = '0.0'
        ws.cell(row=row, column=3).font = Font(bold=True, size=14)

        # ── Benefits by Lever Group ──
        row += 4
        ws[f'B{row}'] = "BENEFITS BY LEVER (Annual Steady State)"
        ws[f'B{row}'].font = Font(bold=True, size=14, color=self.COLORS['primary'])
        row += 2
        self._write_header_row(ws, row, ['Lever Group', 'Lifecycle Stage', 'Annual Benefit'])
        row += 1

        group_total_refs = []
        for group_key in self.group_order:
            group = self.lever_groups[group_key]
            lever_cells = self.cell_map.get('lever_sheets', {}).get(group_key, {})
            annual_cell = lever_cells.get('annual_total')

            ws.cell(row=row, column=2, value=group.get('group_name', group_key))
            ws.cell(row=row, column=3, value=group.get('lifecycle_stage', ''))
            if annual_cell:
                ws.cell(row=row, column=4, value=f"='Journey Analysis'!{annual_cell}")
                group_total_refs.append(f'D{row}')
            ws.cell(row=row, column=4).number_format = '$#,##0'
            ws.cell(row=row, column=4).font = Font(bold=True)
            row += 1

        ws.cell(row=row, column=2, value="TOTAL")
        ws.cell(row=row, column=2).font = Font(bold=True)
        if group_total_refs:
            ws.cell(row=row, column=4, value=f"=SUM({group_total_refs[0]}:{group_total_refs[-1]})")
        ws.cell(row=row, column=4).number_format = '$#,##0'
        ws.cell(row=row, column=4).font = Font(bold=True, size=14, color=self.COLORS['positive'])

    # ── Bank Profile ─────────────────────────────────────────────────

    def _create_bank_profile_sheet(self):
        """Create Bank Profile sheet with identity, key metrics, and data sources."""
        ws = self.wb.create_sheet("Bank Profile", self.sheet_index)
        self.sheet_index += 1

        for col, w in [('A', 5), ('B', 35), ('C', 22), ('D', 12), ('E', 55)]:
            ws.column_dimensions[col].width = w

        profile = self.config.get('bank_profile', {})

        ws['B1'] = "Bank Profile"
        ws['B1'].font = Font(bold=True, size=18, color=self.COLORS['primary'])

        # ── Bank Identity ──
        row = 3
        ws[f'B{row}'] = "BANK IDENTITY"
        ws[f'B{row}'].font = Font(bold=True, size=12, color=self.COLORS['primary'])
        row += 1

        identity_fields = [
            ('Bank Name', profile.get('bank_name', self.config.get('client_name', ''))),
            ('Country', profile.get('country', '')),
            ('Currency', profile.get('currency', self.config.get('currency', ''))),
            ('Stock Ticker', profile.get('stock_ticker', 'N/A')),
            ('Report Year', profile.get('report_year', '')),
            ('Data Sources', profile.get('report_source', '')),
        ]
        for label, value in identity_fields:
            ws.cell(row=row, column=2, value=label)
            ws.cell(row=row, column=2).font = Font(bold=True)
            ws.cell(row=row, column=3, value=str(value) if value else '')
            row += 1

        # Market Context Status
        row += 1
        mc_status = profile.get('market_context_status', 'NOT_RUN')
        ws.cell(row=row, column=2, value="Market Context Status")
        ws.cell(row=row, column=2).font = Font(bold=True)
        ws.cell(row=row, column=3, value=mc_status)
        if mc_status == 'VALIDATED':
            ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor=self.COLORS['transcript'])
        elif mc_status == 'SKIPPED':
            ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor=self.COLORS['benchmark'])
        else:
            ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor=self.COLORS['estimate'])
        row += 2

        # ── Key Financial Metrics ──
        ws[f'B{row}'] = "KEY FINANCIAL METRICS"
        ws[f'B{row}'].font = Font(bold=True, size=12, color=self.COLORS['primary'])
        row += 1
        self._write_header_row(ws, row, ['Metric', 'Value', 'Confidence', 'Source'])
        row += 1

        for metric in profile.get('key_metrics', []):
            ws.cell(row=row, column=2, value=metric.get('metric', ''))
            val = metric.get('value', '')
            ws.cell(row=row, column=3, value=val)
            if isinstance(val, (int, float)) and val > 1000:
                ws.cell(row=row, column=3).number_format = '#,##0'
            elif isinstance(val, float) and 0 < val < 1:
                ws.cell(row=row, column=3).number_format = '0.0%'
            conf = metric.get('confidence', 'LOW')
            ws.cell(row=row, column=4, value=conf)
            if conf == 'HIGH':
                ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor=self.COLORS['transcript'])
            elif conf == 'MEDIUM':
                ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor=self.COLORS['benchmark'])
            else:
                ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor=self.COLORS['estimate'])
            ws.cell(row=row, column=5, value=metric.get('source', ''))
            ws.cell(row=row, column=5).font = Font(size=9)
            row += 1

        # ── Derived Metrics ──
        derived = profile.get('derived_metrics', [])
        if derived:
            row += 1
            ws[f'B{row}'] = "DERIVED METRICS"
            ws[f'B{row}'].font = Font(bold=True, size=12, color=self.COLORS['primary'])
            row += 1
            self._write_header_row(ws, row, ['Metric', 'Value', 'Confidence', 'Derivation'])
            row += 1
            for metric in derived:
                ws.cell(row=row, column=2, value=metric.get('metric', ''))
                val = metric.get('value', '')
                ws.cell(row=row, column=3, value=val)
                if isinstance(val, (int, float)) and val > 1000:
                    ws.cell(row=row, column=3).number_format = '#,##0'
                elif isinstance(val, float) and 0 < val < 1:
                    ws.cell(row=row, column=3).number_format = '0.0%'
                ws.cell(row=row, column=4, value=metric.get('confidence', 'LOW'))
                ws.cell(row=row, column=5, value=metric.get('derivation', ''))
                ws.cell(row=row, column=5).font = Font(size=9)
                row += 1

        # ── Additional Context (Annual Report / Market Research) ──
        additional = profile.get('additional_context', [])
        if additional:
            row += 1
            ws[f'B{row}'] = "ADDITIONAL CONTEXT (Annual Report / Market Research)"
            ws[f'B{row}'].font = Font(bold=True, size=12, color=self.COLORS['primary'])
            row += 1
            self._write_header_row(ws, row, ['Metric', 'Value', 'Confidence', 'Source'])
            row += 1
            for metric in additional:
                ws.cell(row=row, column=2, value=metric.get('metric', ''))
                val = metric.get('value', '')
                ws.cell(row=row, column=3, value=val)
                if isinstance(val, (int, float)) and val > 1000:
                    ws.cell(row=row, column=3).number_format = '#,##0'
                elif isinstance(val, float) and 0 < val < 1:
                    ws.cell(row=row, column=3).number_format = '0.0%'
                conf = metric.get('confidence', 'LOW')
                ws.cell(row=row, column=4, value=conf)
                if conf == 'HIGH':
                    ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor=self.COLORS['transcript'])
                elif conf == 'MEDIUM':
                    ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor=self.COLORS['benchmark'])
                else:
                    ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor=self.COLORS['estimate'])
                ws.cell(row=row, column=5, value=metric.get('source', ''))
                ws.cell(row=row, column=5).font = Font(size=9)
                row += 1

        # ── Data Gaps ──
        gaps = profile.get('data_gaps', [])
        if gaps:
            row += 1
            ws[f'B{row}'] = "DATA GAPS"
            ws[f'B{row}'].font = Font(bold=True, size=12, color=self.COLORS['primary'])
            row += 1
            self._write_header_row(ws, row, ['Data Needed', 'Priority', 'Impact', 'Where to Obtain'])
            row += 1
            for gap in gaps:
                ws.cell(row=row, column=2, value=gap.get('data_needed', ''))
                priority = gap.get('priority', 'MEDIUM')
                ws.cell(row=row, column=3, value=priority)
                if priority in ['CRITICAL', 'HIGH']:
                    ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor=self.COLORS['low_confidence'])
                    ws.cell(row=row, column=3).font = Font(bold=True)
                ws.cell(row=row, column=4, value=gap.get('impact', ''))
                ws.cell(row=row, column=4).font = Font(size=9)
                ws.cell(row=row, column=5, value=gap.get('where_to_obtain', ''))
                ws.cell(row=row, column=5).font = Font(size=9)
                row += 1

    # ── Assumptions (unchanged) ───────────────────────────────────────

    def _create_assumptions_sheet(self):
        ws = self.wb.create_sheet("Assumptions", self.sheet_index)
        self.sheet_index += 1

        for col, w in [('A', 5), ('B', 8), ('C', 35), ('D', 15), ('E', 12), ('F', 50), ('G', 20), ('H', 35)]:
            ws.column_dimensions[col].width = w

        ws['B1'] = "Assumptions Register"
        ws['B1'].font = Font(bold=True, size=18, color=self.COLORS['primary'])
        ws['B3'] = "All assumptions documented with sources and validation owners"
        ws['B3'].font = Font(italic=True)

        row = 5
        self._write_header_row(ws, row, ['ID', 'Assumption', 'Value', 'Conf.', 'Source', 'Owner', 'Sensitivity'])
        row += 1

        assumptions = self.config.get('assumptions_register', None) or self.config.get('assumptions', [])
        for assumption in assumptions:
            ws.cell(row=row, column=2, value=assumption.get('id', ''))
            ws.cell(row=row, column=3, value=assumption.get('assumption', ''))
            ws.cell(row=row, column=4, value=assumption.get('value', ''))

            confidence = assumption.get('confidence', 'LOW')
            ws.cell(row=row, column=5, value=confidence)
            if confidence == 'HIGH':
                ws.cell(row=row, column=4).fill = PatternFill("solid", fgColor=self.COLORS['transcript'])
            elif confidence == 'MEDIUM':
                ws.cell(row=row, column=4).fill = PatternFill("solid", fgColor=self.COLORS['benchmark'])
            else:
                ws.cell(row=row, column=4).fill = PatternFill("solid", fgColor=self.COLORS['estimate'])

            ws.cell(row=row, column=6, value=assumption.get('source', ''))
            ws.cell(row=row, column=6).font = Font(size=9)
            ws.cell(row=row, column=7, value=assumption.get('validation_owner', ''))
            ws.cell(row=row, column=8, value=assumption.get('sensitivity', ''))
            ws.cell(row=row, column=8).font = Font(size=9)
            row += 1

    # ── Data Gaps (unchanged) ─────────────────────────────────────────

    def _create_data_gaps_sheet(self):
        ws = self.wb.create_sheet("Data Gaps", self.sheet_index)
        self.sheet_index += 1

        for col, w in [('A', 5), ('B', 8), ('C', 40), ('D', 12), ('E', 35), ('F', 45)]:
            ws.column_dimensions[col].width = w

        ws['B1'] = "Data Gaps — Required for Validation"
        ws['B1'].font = Font(bold=True, size=18, color=self.COLORS['primary'])
        ws['B3'] = "These data points are CRITICAL for validating the ROI model."
        ws['B3'].font = Font(italic=True, color=self.COLORS['negative'])

        row = 5
        self._write_header_row(ws, row, ['ID', 'Data Needed', 'Priority', 'Source', 'Impact on Model'])
        row += 1

        gaps = self.config.get('data_gaps_for_validation', None) or self.config.get('data_gaps', [])
        for idx, gap in enumerate(gaps, 1):
            # Normalize sub-field names: 'item' → 'data_needed', 'mitigation'/'where_to_obtain' → 'source'
            gap_id = gap.get('id', f'DG-{idx:03d}')
            data_needed = gap.get('data_needed', '') or gap.get('item', '')
            source = gap.get('source', '') or gap.get('mitigation', '') or gap.get('where_to_obtain', '')

            ws.cell(row=row, column=2, value=gap_id)
            ws.cell(row=row, column=3, value=data_needed)
            priority = gap.get('priority', 'MEDIUM')
            ws.cell(row=row, column=4, value=priority)
            if priority == 'CRITICAL':
                ws.cell(row=row, column=4).fill = PatternFill("solid", fgColor=self.COLORS['low_confidence'])
                ws.cell(row=row, column=4).font = Font(bold=True)
            elif priority == 'HIGH':
                ws.cell(row=row, column=4).fill = PatternFill("solid", fgColor=self.COLORS['estimate'])
            ws.cell(row=row, column=5, value=source)
            ws.cell(row=row, column=6, value=gap.get('impact', ''))
            ws.cell(row=row, column=6).font = Font(size=9)
            row += 1

    # ── Named Ranges ──────────────────────────────────────────────────

    def _define_named_ranges(self):
        """Define workbook-level named ranges for key cells."""
        mi = self.cell_map.get('model_inputs', {})

        cf_scenario_row = self.cell_map.get('cashflows', {}).get('scenario_row', 5)
        ranges = {
            'ScenarioIndex': f"'Cashflows'!$C${cf_scenario_row}",
            'DiscountRate': f"'Model Inputs'!${mi.get('discount_rate', 'C20')}",
        }

        # Investment named ranges
        for yr in range(1, 6):
            lic_ref = mi.get('investment', {}).get(f'license_y{yr}')
            if lic_ref:
                ranges[f'License_Y{yr}'] = f"'Model Inputs'!${lic_ref}"
            impl_ref = mi.get('investment', {}).get(f'impl_y{yr}')
            if impl_ref:
                ranges[f'Impl_Y{yr}'] = f"'Model Inputs'!${impl_ref}"

        for name, ref in ranges.items():
            try:
                dn = DefinedName(name, attr_text=ref)
                self.wb.defined_names.add(dn)
            except Exception:
                pass  # Skip if name already exists


def main():
    parser = argparse.ArgumentParser(description='Generate formula-based ROI Excel models')
    parser.add_argument('--output', '-o', default='roi_model.xlsx', help='Output Excel file path')
    parser.add_argument('--config', '-c', required=True, help='JSON configuration file')
    args = parser.parse_args()

    if not Path(args.config).exists():
        print(f"Config file not found: {args.config}")
        exit(1)

    with open(args.config) as f:
        config = json.load(f)

    generator = ROIModelGenerator(config)
    generator.generate(args.output)


if __name__ == '__main__':
    main()
