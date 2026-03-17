"""Excel output renderer using openpyxl."""

from __future__ import annotations

from pathlib import Path

from ..models.output import DealPackage


def render_excel(package: DealPackage, output_path: Path) -> Path:
    """Export deal package financials to an Excel workbook.

    Creates sheets:
    - Summary: KPI overview
    - Recommended: Year-by-year recommended scenario
    - Alternative: Year-by-year alternative (if present)
    - Baseline: Year-by-year baseline

    Args:
        package: Complete deal package.
        output_path: Path where the .xlsx will be written.

    Returns:
        Path to the generated Excel file.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise RuntimeError(
            "openpyxl is required for Excel export. "
            "Install it with: pip install openpyxl"
        )

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1A1F71", end_color="1A1F71", fill_type="solid")
    bold_font = Font(bold=True)
    money_fmt = '#,##0'
    pct_fmt = '0.0%'

    # ── Summary sheet ───────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"

    rec = package.financials.recommended
    base = package.financials.baseline
    fin = package.financials

    summary_data = [
        ("Client", package.deal_brief.client.name),
        ("Deal ID", package.deal_brief.deal_id),
        ("Deal Type", package.deal_brief.deal_type.value),
        ("Segments", ", ".join(s.value for s in package.deal_brief.client.segments)),
        ("", ""),
        ("Baseline ARR", base.arr_year_1),
        ("Recommended ARR (Y1)", rec.arr_year_1),
        ("Recommended ARR (Y5)", rec.arr_year_5),
        ("ARR Uplift %", fin.arr_uplift_pct / 100),
        ("ARR Uplift $", fin.arr_uplift_abs),
        ("5-Year Deal Value", rec.total_deal_value_5yr),
        ("5-Year Uplift vs Baseline", fin.total_uplift_abs),
        ("Avg $/User (5yr)", rec.avg_per_user_5yr),
        ("ARR CAGR", rec.arr_cagr),
        ("", ""),
        ("Recommended Construct", rec.construct_type.value),
        ("Rationale", package.pricing_strategy.rationale),
    ]

    for row_idx, (label, value) in enumerate(summary_data, 1):
        ws.cell(row=row_idx, column=1, value=label).font = bold_font
        cell = ws.cell(row=row_idx, column=2, value=value)
        if isinstance(value, float) and abs(value) >= 1000:
            cell.number_format = money_fmt
        elif isinstance(value, float) and abs(value) < 10:
            cell.number_format = pct_fmt

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 40

    # ── Yearly sheets ───────────────────────────────────────
    def _write_yearly_sheet(ws, model, title):
        headers = ["Year", "Users", "DB Revenue", "Assist Revenue", "Lending Revenue",
                    "Add-on Revenue", "Total License", "Services", "Total Revenue", "$/User", "ARR"]

        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, y in enumerate(model.yearly, 2):
            ws.cell(row=row_idx, column=1, value=f"Y{y.year}")
            ws.cell(row=row_idx, column=2, value=y.users).number_format = '#,##0'
            ws.cell(row=row_idx, column=3, value=y.digital_banking_revenue).number_format = money_fmt
            ws.cell(row=row_idx, column=4, value=y.assist_revenue).number_format = money_fmt
            ws.cell(row=row_idx, column=5, value=y.lending_revenue).number_format = money_fmt
            ws.cell(row=row_idx, column=6, value=y.add_on_revenue).number_format = money_fmt
            ws.cell(row=row_idx, column=7, value=y.license_revenue).number_format = money_fmt
            ws.cell(row=row_idx, column=8, value=y.services_revenue).number_format = money_fmt
            ws.cell(row=row_idx, column=9, value=y.total_revenue).number_format = money_fmt
            ws.cell(row=row_idx, column=10, value=y.per_user_cost).number_format = '#,##0.00'
            ws.cell(row=row_idx, column=11, value=y.arr).number_format = money_fmt

        # Totals row
        total_row = len(model.yearly) + 2
        ws.cell(row=total_row, column=1, value="Total").font = bold_font
        ws.cell(row=total_row, column=7, value=model.total_license_5yr).number_format = money_fmt
        ws.cell(row=total_row, column=7).font = bold_font
        ws.cell(row=total_row, column=8, value=model.total_services_5yr).number_format = money_fmt
        ws.cell(row=total_row, column=8).font = bold_font
        ws.cell(row=total_row, column=9, value=model.total_deal_value_5yr).number_format = money_fmt
        ws.cell(row=total_row, column=9).font = bold_font

        # Auto-width
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 16

    _write_yearly_sheet(
        wb.create_sheet("Recommended"),
        package.financials.recommended,
        "Recommended",
    )

    if package.financials.alternative:
        _write_yearly_sheet(
            wb.create_sheet("Alternative"),
            package.financials.alternative,
            "Alternative",
        )

    _write_yearly_sheet(
        wb.create_sheet("Baseline"),
        package.financials.baseline,
        "Baseline",
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path
