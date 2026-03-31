#!/usr/bin/env python3
"""AI Pricing Stress Report — generates Excel workbook with scale × domain analysis.

Usage:
    python tools/ai_pricing_stress_report.py [output_path]

Produces a workbook with:
- One sheet per domain (retail, wealth, commercial)
- Rows: user scale points (50, 100, 225, 500, 1000, 2000, 5000)
- Columns: all pricing layers, margins, per-user costs
- Summary sheet with cross-domain comparison
"""

import sys
from pathlib import Path

# Add src to path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

from deal_pricing.engine.ai_pricing_engine import compute_ai_pricing, load_agent_catalog
from deal_pricing.models.ai_pricing import AIPricingConfig, AIPillar

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl required: pip install openpyxl")
    sys.exit(1)


SCALE_POINTS = [50, 100, 225, 500, 1000, 2000, 5000]
DOMAINS = ["retail", "wealth", "commercial"]
TIERS = {"retail": "entry", "wealth": "standard", "commercial": "entry"}

# Styles
HEADER_FILL = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
GOOD_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
BAD_FILL = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
BORDER = Border(
    bottom=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
)


def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def format_currency(cell):
    cell.number_format = '#,##0'


def format_pct(cell):
    cell.number_format = '0.0'


def write_domain_sheet(wb, domain):
    ws = wb.create_sheet(title=domain.capitalize())
    tier = TIERS.get(domain, "entry")
    catalog = load_agent_catalog(domain)
    agent_count = len(catalog)

    headers = [
        "Users", "Agents", "IF Tier",
        "Raw Compute ($)", "Platform Price ($)", "Build-It-Yourself ($)",
        "Platform:Compute", "Per User/Year ($)", "Per User/Month ($)",
        "Annual BICs", "BIC Blocks",
        "Domain Base ($)", "BIC Blocks ($)", "Compute P/T ($)",
        "Margin vs BIY (%)",
    ]

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"AI Pricing Stress Analysis — {domain.capitalize()} Banking"
    title_cell.font = Font(name="Calibri", size=14, bold=True)

    ws.cell(row=2, column=1).value = f"Agent catalog: {agent_count} agents | IF Domain Tier: {tier}"
    ws.cell(row=2, column=1).font = Font(name="Calibri", size=10, italic=True, color="666666")

    # Headers
    header_row = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=i).value = h
    style_header(ws, header_row, len(headers))

    # Data rows
    for row_idx, users in enumerate(SCALE_POINTS, header_row + 1):
        config = AIPricingConfig(
            pillars=[AIPillar.EMBEDDED_AI, AIPillar.PROCESS_AUTOMATION],
            if_domain_tier=tier,
        )
        r = compute_ai_pricing(config, user_count=users, domain=domain)
        ifp = r.if_pricing

        margin_vs_biy = ((r.build_it_yourself_annual_usd - r.platform_price_annual_usd)
                         / r.build_it_yourself_annual_usd * 100)

        values = [
            users, r.total_agents, tier,
            r.raw_compute_annual_usd, r.platform_price_annual_usd, r.build_it_yourself_annual_usd,
            f"{r.platform_to_compute_ratio}x", r.per_user_ai_cost_annual_usd, r.per_user_ai_cost_monthly_usd,
            r.bic_consumption.total_annual_bics, r.bic_consumption.blocks_required,
            ifp.domain_base_annual_usd, ifp.bic_blocks_annual_usd, ifp.compute_passthrough_annual_usd,
            round(margin_vs_biy, 1),
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.border = BORDER

            # Currency formatting
            if col_idx in (4, 5, 6, 8, 9, 12, 13, 14):
                format_currency(cell)
            elif col_idx == 15:
                format_pct(cell)

            # Conditional formatting for platform:compute ratio
            if col_idx == 7:
                cell.alignment = Alignment(horizontal="center")

    # Column widths
    widths = [8, 8, 10, 16, 16, 18, 14, 14, 14, 14, 10, 14, 14, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return ws


def write_summary_sheet(wb):
    ws = wb.create_sheet(title="Summary")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws.cell(row=1, column=1).value = "Cross-Domain AI Pricing Summary"
    ws.cell(row=1, column=1).font = Font(name="Calibri", size=14, bold=True)

    headers = ["Domain", "Users", "Platform ($)", "Per User/Mo ($)", "BICs/yr", "Ratio", "BIY ($)", "Saving vs BIY"]
    header_row = 3
    for i, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=i).value = h
    style_header(ws, header_row, len(headers))

    row = header_row + 1
    for domain in DOMAINS:
        tier = TIERS.get(domain, "entry")
        for users in [50, 225, 1000, 5000]:
            config = AIPricingConfig(
                pillars=[AIPillar.EMBEDDED_AI, AIPillar.PROCESS_AUTOMATION],
                if_domain_tier=tier,
            )
            r = compute_ai_pricing(config, user_count=users, domain=domain)
            saving = r.build_it_yourself_annual_usd - r.platform_price_annual_usd

            values = [
                domain.capitalize(), users,
                r.platform_price_annual_usd,
                r.per_user_ai_cost_monthly_usd,
                r.bic_consumption.total_annual_bics,
                f"{r.platform_to_compute_ratio}x",
                r.build_it_yourself_annual_usd,
                saving,
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = val
                cell.border = BORDER
                if col in (3, 4, 7, 8):
                    format_currency(cell)

            # Color code the saving
            saving_cell = ws.cell(row=row, column=8)
            if saving > 1000000:
                saving_cell.fill = GOOD_FILL
            elif saving > 500000:
                saving_cell.fill = WARN_FILL

            row += 1
        row += 1  # gap between domains

    for i, w in enumerate([12, 8, 14, 14, 14, 10, 14, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main():
    output_path = sys.argv[1] if len(sys.argv) > 1 else "ai_pricing_stress_report.xlsx"

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    for domain in DOMAINS:
        write_domain_sheet(wb, domain)

    write_summary_sheet(wb)

    wb.save(output_path)
    print(f"Stress report saved to {output_path}")

    # Print summary to console
    print("\n=== Quick Summary ===\n")
    for domain in DOMAINS:
        tier = TIERS.get(domain, "entry")
        for users in [225, 1000, 5000]:
            config = AIPricingConfig(
                pillars=[AIPillar.EMBEDDED_AI, AIPillar.PROCESS_AUTOMATION],
                if_domain_tier=tier,
            )
            r = compute_ai_pricing(config, user_count=users, domain=domain)
            print(f"{domain:12s} | {users:5d} users | "
                  f"Platform: ${r.platform_price_annual_usd:>10,.0f} | "
                  f"Per user: ${r.per_user_ai_cost_monthly_usd:>6,.0f}/mo | "
                  f"Ratio: {r.platform_to_compute_ratio:>5.1f}x | "
                  f"BIY: ${r.build_it_yourself_annual_usd:>12,.0f}")
        print()


if __name__ == "__main__":
    main()
